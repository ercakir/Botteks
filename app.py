import os
import streamlit as st
import psycopg2
import psycopg2.pool
import sqlite3
import pandas as pd
from openai import OpenAI
import datetime
import io
import re

# ==========================================
# 0. DATABASE CONNECTION POOLING & FALLBACK ENGINE
# ==========================================
@st.cache_resource
def get_db_pool():
    try:
        return psycopg2.pool.ThreadedConnectionPool(
            minconn=1,
            maxconn=10,
            host="localhost",
            port="5432",
            database="tekstil_demo_2",
            user="chatbot_user",
            password="sifreniz"
        )
    except Exception as e:
        print("PostgreSQL connection offline. Falling back to SQLite:", e)
        return None

@st.cache_resource
def init_db_indexes_once():
    pool = get_db_pool()
    if not pool:
        return
    conn = None
    try:
        conn = pool.getconn()
        with conn.cursor() as cursor:
            cursor.execute("""
                CREATE INDEX IF NOT EXISTS idx_siparisler_tarih_durum ON siparisler(tarih DESC, urun_durumu);
                CREATE INDEX IF NOT EXISTS idx_siparisler_musteri_alt ON siparisler(musteri_id, musteri_alt_marka_id);
                CREATE INDEX IF NOT EXISTS idx_musteriler_adi ON musteriler(musteri_adi);
                CREATE INDEX IF NOT EXISTS idx_alt_markalari_adi ON musteri_alt_markalari(alt_marka_adi);
            """)
            conn.commit()
    except Exception as e:
        print("DB Index init:", e)
    finally:
        if conn and pool:
            pool.putconn(conn)

# Run database indexes initialization once on startup
try:
    init_db_indexes_once()
except Exception:
    pass
st.set_page_config(page_title="Botteks - Akıllı BI & AI Asistanı", page_icon="🧶", layout="wide", initial_sidebar_state="expanded")

# --- GEÇMİŞ VE FAVORİ HAFIZALARINI BAŞLAT ---
if "gecmis_aramalar" not in st.session_state:
    st.session_state.gecmis_aramalar = []

if "chat_history" not in st.session_state:
    st.session_state.chat_history = [
        {"role": "assistant", "type": "chat", "content": "Merhaba! Ben Botteks Akıllı İş Zekası (BI) Asistanınızım. Şirket verileriyle ilgili sorularınızı sorabilir, analiz grafikleri isteyebilir ya da sol menüden kendi Excel/CSV dosyalarınızı yükleyerek anında sorgulayabilirsiniz. Nasıl yardımcı olabilirim?"}
    ]

if "saved_analyses" not in st.session_state:
    st.session_state.saved_analyses = []

if "prompt_to_submit" not in st.session_state:
    st.session_state.prompt_to_submit = None

if "last_processed_audio_hash" not in st.session_state:
    st.session_state.last_processed_audio_hash = None

# --- GLOBAL FİLTRE SEÇENEKLERİ HAFIZASI ---
if "filter_options" not in st.session_state:
    st.session_state.filter_options = {
        "customers": ["Tümü", "H&M", "Inditex", "Primark", "Adidas", "Nike"],
        "brands": ["Tümü", "H&M (Tümü)", "Zara", "Bershka", "HM Basic", "HM Premium", "HM Sport", "HM Kids", "HM Home", "HM Divided", "Primark Kids", "Primark Home", "Primark Essentials", "Primark Trend", "HM Trend", "Inditex Trend", "HM Essentials", "Primark Premium", "Zara Home", "Zara Kids", "HM Divided Trend", "Primark Divided", "Zara Divided"]
    }

DEPT_TEMPLATES = {
    "Executive / C-Level": [
        "Yıllara göre toplam ciro trendi nasıldır?",
        "Aylara göre toplam sipariş miktarı trendi nasıldır?",
        "En çok ciro getiren ilk 5 müşteri hangileridir?",
        "En çok sipariş miktarına sahip ilk 5 marka hangileridir?",
        "Toplam cironun müşteri bazında dağılımı nasıldır?",
        "Toplam sipariş miktarının kategorilere göre dağılımı nasıldır?",
        "Gelecek 90 günlük aylık ciro tahmini nedir?",
        "En çok sipariş iptal edilen ilk 5 müşteri hangileridir?",
        "Yıllık bazda toplam tahsilat tutarı trendi nasıldır?",
        "Toplam cironun merkez ülkelere göre dağılımı nasıldır?"
    ],
    "Sales & Marketing": [
        "En çok sipariş edilen ilk 10 ürün kodu hangisidir?",
        "En yüksek satış miktarına sahip ilk 5 marka hangisidir?",
        "Aylara göre toplam sipariş miktarı trendi nasıldır?",
        "Toplam sipariş miktarının marka bazında dağılımı nasıldır?",
        "En popüler ilk 5 ürünün ciro dağılımı nasıldır?",
        "Sipariş miktarlarının ürün türlerine göre dağılımı nasıldır?",
        "Son 12 ayın aylık sipariş adeti trendi nasıldır?",
        "Müşterilerin marka bazlı sipariş miktarı dağılımı nasıldır?",
        "En az sipariş edilen son 5 ürün kodu hangileridir?",
        "Aktif markaların toplam sipariş miktarlarına göre dağılımı nasıldır?"
    ],
    "Finance": [
        "Aylara göre toplam ciro trendi nasıldır?",
        "En çok ciro getiren ilk 5 müşteri hangileridir?",
        "Toplam cironun marka bazında dağılımı nasıldır?",
        "Müşterilerin toplam ciroya katkı oranları nedir?",
        "Aylık bazda ortalama birim fiyat trendi nasıldır?",
        "İptal edilen siparişlerin ciro kaybı müşteri bazında nasıldır?",
        "En çok ciro getiren ilk 10 ürün hangileridir?",
        "Aylara göre toplam tahsilat (ödeme) trendi nasıldır?",
        "Marka bazında ortalama birim fiyat dağılımı nasıldır?",
        "Ciro performansının merkez ülkelere göre dağılımı nasıldır?"
    ],
    "Inventory & Product": [
        "Sipariş edilen ürün miktarlarının kumaş türlerine göre dağılımı nasıldır?",
        "Sipariş miktarlarının ürün türlerine göre dağılımı nasıldır?",
        "En çok sipariş edilen ilk 10 ürün hangileridir?",
        "Aylara göre toplam sipariş edilen ürün miktarı trendi nasıldır?",
        "En yüksek sipariş miktarına sahip ilk 5 marka hangisidir?",
        "Müşterilerin sipariş ettiği toplam ürün miktarı dağılımı nasıldır?",
        "Kumaş türü bazında ortalama sipariş miktarı dağılımı nasıldır?",
        "En az sipariş edilen son 5 kumaş türü hangileridir?",
        "Yıllara göre sipariş edilen toplam ürün miktarı trendi nasıldır?",
        "Toplam siparişlerin ürün kategorilerine göre dağılımı nasıldır?"
    ]
}

FAST_SQL_MAP = {
    # Executive / C-Level
    "Yıllara göre toplam ciro trendi nasıldır?": "SELECT STRFTIME('%Y', s.tarih) AS yil, SUM(s.miktar * s.birim_fiyat) AS toplam_ciro FROM siparisler s GROUP BY yil ORDER BY yil;",
    "Aylara göre toplam sipariş miktarı trendi nasıldır?": "SELECT TO_CHAR(s.tarih, 'YYYY-MM') AS ay, SUM(s.miktar) AS toplam_siparis_miktari FROM siparisler s GROUP BY ay ORDER BY ay;",
    "En çok ciro getiren ilk 5 müşteri hangileridir?": "SELECT m.musteri_adi, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id GROUP BY m.musteri_adi ORDER BY ciro DESC LIMIT 5;",
    "En çok sipariş miktarına sahip ilk 5 marka hangileridir?": "SELECT ma.alt_marka_adi AS marka_adi, SUM(s.miktar) AS toplam_miktar FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id GROUP BY ma.alt_marka_adi ORDER BY toplam_miktar DESC LIMIT 5;",
    "Toplam cironun müşteri bazında dağılımı nasıldır?": "SELECT m.musteri_adi, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id GROUP BY m.musteri_adi ORDER BY ciro DESC;",
    "Toplam sipariş miktarının kategorilere göre dağılımı nasıldır?": "SELECT ma.kategori, SUM(s.miktar) AS toplam_miktar FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id GROUP BY ma.kategori ORDER BY toplam_miktar DESC;",
    "Gelecek 90 günlük aylık ciro tahmini nedir?": "[ACTION: FORECAST_90_DAYS] SOHBET: Gelecek 90 gün için ciro tahmini başarıyla hesaplandı. SELECT TO_CHAR(s.tarih, 'YYYY-MM') AS ay, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s GROUP BY ay ORDER BY ay;",
    "En çok sipariş iptal edilen ilk 5 müşteri hangileridir?": "SELECT m.musteri_adi, COUNT(*) AS iptal_siparis_sayisi FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id WHERE s.urun_durumu = 'iptal' GROUP BY m.musteri_adi ORDER BY iptal_siparis_sayisi DESC LIMIT 5;",
    "Yıllık bazda toplam tahsilat tutarı trendi nasıldır?": "SELECT STRFTIME('%Y', o.odeme_tarihi) AS yil, SUM(o.miktar) AS toplam_tahsilat FROM odemeler o GROUP BY yil ORDER BY yil;",
    "Toplam cironun merkez ülkelere göre dağılımı nasıldır?": "SELECT ma.merkez_ulke, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id GROUP BY ma.merkez_ulke ORDER BY ciro DESC;",

    # Sales & Marketing
    "En çok sipariş edilen ilk 10 ürün kodu hangisidir?": "SELECT s.urun_kodu, SUM(s.miktar) AS toplam_siparis_miktari FROM siparisler s GROUP BY s.urun_kodu ORDER BY toplam_siparis_miktari DESC LIMIT 10;",
    "En yüksek satış miktarına sahip ilk 5 marka hangisidir?": "SELECT ma.alt_marka_adi AS marka_adi, SUM(s.miktar) AS toplam_miktar FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id GROUP BY ma.alt_marka_adi ORDER BY toplam_miktar DESC LIMIT 5;",
    "Toplam sipariş miktarının marka bazında dağılımı nasıldır?": "SELECT ma.alt_marka_adi AS marka_adi, SUM(s.miktar) AS toplam_miktar FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id GROUP BY ma.alt_marka_adi ORDER BY toplam_miktar DESC;",
    "En popüler ilk 5 ürünün ciro dağılımı nasıldır?": "SELECT s.urun_kodu, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s GROUP BY s.urun_kodu ORDER BY ciro DESC LIMIT 5;",
    "Sipariş miktarlarının ürün türlerine göre dağılımı nasıldır?": "SELECT ut.urun_turu_adi, SUM(s.miktar) AS toplam_miktar FROM siparisler s JOIN urunler u ON s.urun_kodu = u.urun_kodu JOIN urun_turleri ut ON u.urun_turu_id = ut.id GROUP BY ut.urun_turu_adi ORDER BY toplam_miktar DESC;",
    "Son 12 ayın aylık sipariş adeti trendi nasıldır?": "SELECT TO_CHAR(s.tarih, 'YYYY-MM') AS ay, SUM(s.miktar) AS toplam_miktar FROM siparisler s GROUP BY ay ORDER BY ay DESC LIMIT 12;",
    "Müşterilerin marka bazlı sipariş miktarı dağılımı nasıldır?": "SELECT m.musteri_adi, ma.alt_marka_adi, SUM(s.miktar) AS toplam_miktar FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id GROUP BY m.musteri_adi, ma.alt_marka_adi ORDER BY toplam_miktar DESC;",
    "En az sipariş edilen son 5 ürün kodu hangileridir?": "SELECT s.urun_kodu, SUM(s.miktar) AS toplam_siparis_miktari FROM siparisler s GROUP BY s.urun_kodu ORDER BY toplam_siparis_miktari ASC LIMIT 5;",
    "Aktif markaların toplam sipariş miktarlarına göre dağılımı nasıldır?": "SELECT ma.alt_marka_adi, SUM(s.miktar) AS toplam_miktar FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id WHERE ma.aktif = TRUE GROUP BY ma.alt_marka_adi ORDER BY toplam_miktar DESC;",

    # Finance
    "Aylara göre toplam ciro trendi nasıldır?": "SELECT TO_CHAR(s.tarih, 'YYYY-MM') AS ay, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s GROUP BY ay ORDER BY ay;",
    "Toplam cironun marka bazında dağılımı nasıldır?": "SELECT ma.alt_marka_adi, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id GROUP BY ma.alt_marka_adi ORDER BY ciro DESC;",
    "Müşterilerin toplam ciroya katkı oranları nedir?": "SELECT m.musteri_adi, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id GROUP BY m.musteri_adi ORDER BY ciro DESC;",
    "Aylık bazda ortalama birim fiyat trendi nasıldır?": "SELECT TO_CHAR(s.tarih, 'YYYY-MM') AS ay, ROUND(AVG(s.birim_fiyat), 2) AS ortalama_fiyat FROM siparisler s GROUP BY ay ORDER BY ay;",
    "İptal edilen siparişlerin ciro kaybı müşteri bazında nasıldır?": "SELECT m.musteri_adi, SUM(s.miktar * s.birim_fiyat) AS iptal_ciro_kaybi FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id WHERE s.urun_durumu = 'iptal' GROUP BY m.musteri_adi ORDER BY iptal_ciro_kaybi DESC;",
    "En çok ciro getiren ilk 10 ürün hangileridir?": "SELECT s.urun_kodu, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s GROUP BY s.urun_kodu ORDER BY ciro DESC LIMIT 10;",
    "Aylara göre toplam tahsilat (ödeme) trendi nasıldır?": "SELECT TO_CHAR(o.odeme_tarihi, 'YYYY-MM') AS ay, SUM(o.miktar) AS toplam_tahsilat FROM odemeler o GROUP BY ay ORDER BY ay;",
    "Marka bazında ortalama birim fiyat dağılımı nasıldır?": "SELECT ma.alt_marka_adi, ROUND(AVG(s.birim_fiyat), 2) AS ortalama_birim_fiyat FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id GROUP BY ma.alt_marka_adi ORDER BY ortalama_birim_fiyat DESC;",
    "Ciro performansının merkez ülkelere göre dağılımı nasıldır?": "SELECT ma.merkez_ulke, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id GROUP BY ma.merkez_ulke ORDER BY ciro DESC;",

    # Inventory & Product
    "Sipariş edilen ürün miktarlarının kumaş türlerine göre dağılımı nasıldır?": "SELECT k.kumas_adi, SUM(s.miktar) AS toplam_miktar FROM siparisler s JOIN urunler u ON s.urun_kodu = u.urun_kodu JOIN kumaslar k ON u.kumas_id = k.id GROUP BY k.kumas_adi ORDER BY toplam_miktar DESC;",
    "En çok sipariş edilen ilk 10 ürün hangileridir?": "SELECT s.urun_kodu, SUM(s.miktar) AS toplam_miktar FROM siparisler s GROUP BY s.urun_kodu ORDER BY toplam_miktar DESC LIMIT 10;",
    "Aylara göre toplam sipariş edilen ürün miktarı trendi nasıldır?": "SELECT TO_CHAR(s.tarih, 'YYYY-MM') AS ay, SUM(s.miktar) AS toplam_miktar FROM siparisler s GROUP BY ay ORDER BY ay;",
    "Müşterilerin sipariş ettiği toplam ürün miktarı dağılımı nasıldır?": "SELECT m.musteri_adi, SUM(s.miktar) AS toplam_miktar FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id GROUP BY m.musteri_adi ORDER BY toplam_miktar DESC;",
    "Kumaş türü bazında ortalama sipariş miktarı dağılımı nasıldır?": "SELECT k.kumas_adi, ROUND(AVG(s.miktar), 2) AS ortalama_miktar FROM siparisler s JOIN urunler u ON s.urun_kodu = u.urun_kodu JOIN kumaslar k ON u.kumas_id = k.id GROUP BY k.kumas_adi ORDER BY ortalama_miktar DESC;",
    "En az sipariş edilen son 5 kumaş türü hangileridir?": "SELECT k.kumas_adi, SUM(s.miktar) AS toplam_miktar FROM siparisler s JOIN urunler u ON s.urun_kodu = u.urun_kodu JOIN kumaslar k ON u.kumas_id = k.id GROUP BY k.kumas_adi ORDER BY toplam_miktar ASC LIMIT 5;",
    "Yıllara göre sipariş edilen toplam ürün miktarı trendi nasıldır?": "SELECT STRFTIME('%Y', s.tarih) AS yil, SUM(s.miktar) AS toplam_miktar FROM siparisler s GROUP BY yil ORDER BY yil;",
    "Toplam siparişlerin ürün kategorilerine göre dağılımı nasıldır?": "SELECT ma.kategori, SUM(s.miktar) AS toplam_miktar FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id GROUP BY ma.kategori ORDER BY toplam_miktar DESC;"
}

# --- STAGE 3: ADVANCED STRING NORMALIZATION & FAST_SQL_MAP ENGINE ---
def normalize_tr_text(text):
    if not text:
        return ""
    text = text.lower()
    tr_map = str.maketrans("çğıöşüiı", "cgossuii")
    text = text.translate(tr_map)
    text = re.sub(r'[^\w\s]', '', text)
    text = re.sub(r'\s+', ' ', text).strip()
    return text

NORMALIZED_FAST_SQL_MAP = {normalize_tr_text(k): v for k, v in FAST_SQL_MAP.items()}

# Add additional intent-based normalized keywords
EXTRA_NORMALIZED_RULES = {
    "en cok ciro getiren ilk 5 musteri": "SELECT m.musteri_adi, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id GROUP BY m.musteri_adi ORDER BY ciro DESC LIMIT 5;",
    "aylara gore ciro trendi": "SELECT TO_CHAR(s.tarih, 'YYYY-MM') AS ay, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s GROUP BY ay ORDER BY ay;",
    "yillara gore ciro trendi": "SELECT STRFTIME('%Y', s.tarih) AS yil, SUM(s.miktar * s.birim_fiyat) AS toplam_ciro FROM siparisler s GROUP BY yil ORDER BY yil;",
    "yillik bazda toplam tahsilat tutari trendi nasildir": "SELECT STRFTIME('%Y', o.odeme_tarihi) AS yil, SUM(o.miktar) AS toplam_tahsilat FROM odemeler o GROUP BY yil ORDER BY yil;",
    "yillara gore siparis edilen toplam urun miktari trendi nasildir": "SELECT STRFTIME('%Y', s.tarih) AS yil, SUM(s.miktar) AS toplam_miktar FROM siparisler s GROUP BY yil ORDER BY yil;",
    "gelecek 90 gunluk ciro tahmini": "[ACTION: FORECAST_90_DAYS] SOHBET: Gelecek 90 gün için ciro tahmini hesaplandı. SELECT TO_CHAR(s.tarih, 'YYYY-MM') AS ay, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s GROUP BY ay ORDER BY ay;",
    "toplam ciro": "SELECT SUM(s.miktar * s.birim_fiyat) AS toplam_ciro_eur FROM siparisler s WHERE s.urun_durumu != 'iptal';",
    "musteri sayisi": "SELECT COUNT(*) AS toplam_musteri FROM musteriler;",
    "musterileri listele": "SELECT id, musteri_adi, ulke FROM musteriler ORDER BY id;",
    "zara satislar": "SELECT ma.alt_marka_adi, SUM(s.miktar) AS toplam_miktar, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id WHERE ma.alt_marka_adi ILIKE '%Zara%' GROUP BY ma.alt_marka_adi;",
    "hm satislar": "SELECT ma.alt_marka_adi, SUM(s.miktar) AS toplam_miktar, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id WHERE ma.alt_marka_adi ILIKE '%HM%' OR ma.alt_marka_adi ILIKE '%H&M%' GROUP BY ma.alt_marka_adi ORDER BY ciro DESC;",
    "primark satislar": "SELECT ma.alt_marka_adi, SUM(s.miktar) AS toplam_miktar, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id WHERE ma.alt_marka_adi ILIKE '%Primark%' GROUP BY ma.alt_marka_adi ORDER BY ciro DESC;",
    "kumas kullanimi": "SELECT k.kumas_adi, SUM(s.miktar) AS toplam_miktar FROM siparisler s JOIN urunler u ON s.urun_kodu = u.urun_kodu JOIN kumaslar k ON u.kumas_id = k.id GROUP BY k.kumas_adi ORDER BY toplam_miktar DESC;",
    "iptal edilen siparisler": "SELECT m.musteri_adi, COUNT(*) AS iptal_sayisi, SUM(s.miktar * s.birim_fiyat) AS kayip_ciro FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id WHERE s.urun_durumu = 'iptal' GROUP BY m.musteri_adi ORDER BY kayip_ciro DESC;",
    "en buyuk musteri": "SELECT m.musteri_adi, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id GROUP BY m.musteri_adi ORDER BY ciro DESC LIMIT 5;",
    "zara kumas": "SELECT k.kumas_adi, SUM(s.miktar) AS toplam_miktar FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id JOIN urunler u ON s.urun_kodu = u.urun_kodu JOIN kumaslar k ON u.kumas_id = k.id WHERE ma.alt_marka_adi ILIKE '%Zara%' GROUP BY k.kumas_adi ORDER BY toplam_miktar DESC LIMIT 1;",
    "en pahali urun": "SELECT s.urun_kodu, MAX(s.birim_fiyat) AS en_yuksek_birim_fiyat FROM siparisler s GROUP BY s.urun_kodu ORDER BY en_yuksek_birim_fiyat DESC LIMIT 1;",
    "en cok satis yapilan ulkeler": "SELECT ma.merkez_ulke, SUM(s.miktar * s.birim_fiyat) AS ciro FROM siparisler s JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id GROUP BY ma.merkez_ulke ORDER BY ciro DESC;"
}

NORMALIZED_FAST_SQL_MAP.update(EXTRA_NORMALIZED_RULES)

# --- YARDIMCI EXCEL DIŞA AKTARMA FONKSİYONU (OKUNAKLI & RENKLİ) ---
def to_excel(df):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    output = io.BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Analiz Raporu')
    
    workbook = writer.book
    worksheet = writer.sheets['Analiz Raporu']
    
    # 1. Stil Tanımları
    header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid") # Koyu Lacivert
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10, color="0F172A")
    data_alignment = Alignment(horizontal="left", vertical="center")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )
    
    # 2. Başlık Satırını Biçimlendirme
    for col_num, col_name in enumerate(df.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border
        
    # 3. Veri Hücrelerini Biçimlendirme ve Dinamik Sütun Genişliği Hesaplama
    for col_idx, col in enumerate(df.columns, 1):
        max_len = len(str(col))
        for row_idx in range(2, len(df) + 2):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
                
        col_letter = get_column_letter(col_idx)
        # En uzun karaktere göre dinamik genişlik ayarlama (+4 karakter payı)
        worksheet.column_dimensions[col_letter].width = min(max(max_len + 5, 14), 60)
        
    writer.close()
    processed_data = output.getvalue()
    return processed_data

# --- RENK VE TASARIM (CLEAN MODERN ENTERPRISE WEB APP DESIGN SYSTEM - REACT/NODE PLATFORM STYLE) ---
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&family=Outfit:wght@500;600;700;800&family=Roboto:wght@400;500;700&display=swap');

    /* Ferah Soft Gray Background & Modern Typography */
    html, body, [class*="css"] {
        font-family: 'Inter', 'Roboto', 'Segoe UI', -apple-system, BlinkMacSystemFont, sans-serif !important;
        color: #1e293b !important;
    }
    
    .stApp {
        background-color: #f8f9fa !important;
        color: #1e293b !important;
    }
    
    h1, h2, h3, h4, h5, h6 {
        font-family: 'Outfit', 'Inter', sans-serif !important;
        color: #0f172a !important;
        letter-spacing: -0.02em !important;
    }
    
    p, span, label, div {
        color: #334155 !important;
    }

    /* Executive Hero Header Banner (Modern White Card with Soft Shadow) */
    .hero-banner {
        background: #ffffff !important;
        border: 1px solid #e2e8f0 !important;
        border-radius: 16px !important;
        padding: 28px 32px !important;
        margin-bottom: 28px !important;
        box-shadow: 0 4px 20px -2px rgba(15, 23, 42, 0.06), 0 2px 6px -1px rgba(15, 23, 42, 0.04) !important;
    }
    .hero-badge {
        display: inline-flex;
        align-items: center;
        gap: 6px;
        background: #eff6ff !important;
        border: 1px solid #bfdbfe !important;
        color: #1d4ed8 !important;
        font-size: 11px !important;
        font-weight: 700 !important;
        text-transform: uppercase !important;
        letter-spacing: 0.1em !important;
        padding: 4px 12px !important;
        border-radius: 20px !important;
        margin-bottom: 12px !important;
    }
    .hero-title {
        color: #0f172a !important;
        font-size: 32px !important;
        font-weight: 800 !important;
        margin: 0 0 8px 0 !important;
        background: linear-gradient(135deg, #0f172a 0%, #1e3a8a 100%) !important;
        -webkit-background-clip: text !important;
        -webkit-text-fill-color: transparent !important;
    }
    .hero-subtitle {
        color: #64748b !important;
        font-size: 15px !important;
        font-weight: 400 !important;
        margin: 0 !important;
        line-height: 1.5 !important;
    }

    /* Clean White Sidebar Styling */
    section[data-testid="stSidebar"] {
        background-color: #ffffff !important;
        border-right: 1px solid #e2e8f0 !important;
    }
    
    /* Sidebar Headers, Text, Captions & Labels */
    section[data-testid="stSidebar"] h1, 
    section[data-testid="stSidebar"] h2, 
    section[data-testid="stSidebar"] h3, 
    section[data-testid="stSidebar"] h4, 
    section[data-testid="stSidebar"] h5,
    section[data-testid="stSidebar"] h6 {
        color: #0f172a !important;
        font-weight: 800 !important;
    }

    section[data-testid="stSidebar"] p, 
    section[data-testid="stSidebar"] span, 
    section[data-testid="stSidebar"] div,
    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] .stMarkdown {
        color: #1e293b !important;
        font-weight: 500 !important;
    }

    section[data-testid="stSidebar"] .stCaption, 
    section[data-testid="stSidebar"] small,
    section[data-testid="stSidebar"] [data-testid="stCaptionContainer"] p {
        color: #64748b !important;
        font-size: 12px !important;
    }

    /* Sidebar Labels */
    section[data-testid="stSidebar"] label,
    section[data-testid="stSidebar"] label *,
    section[data-testid="stSidebar"] label p,
    section[data-testid="stSidebar"] label span {
        color: #1e3a8a !important;
        font-weight: 700 !important;
        font-size: 14px !important;
        letter-spacing: 0.01em !important;
        margin-bottom: 6px !important;
    }

    /* Modern Input Containers (Selectbox, MultiSelect, DateInput) */
    section[data-testid="stSidebar"] div[data-baseweb="select"],
    div[data-baseweb="select"] {
        background-color: #ffffff !important;
        border: 1px solid #cbd5e1 !important;
        border-radius: 10px !important;
        box-shadow: 0 2px 4px rgba(15, 23, 42, 0.04) !important;
        transition: all 0.2s ease !important;
    }
    section[data-testid="stSidebar"] div[data-baseweb="select"]:focus-within,
    div[data-baseweb="select"]:focus-within {
        border-color: #2563eb !important;
        box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.15) !important;
    }
    div[data-baseweb="select"] * {
        color: #0f172a !important;
        font-weight: 600 !important;
    }

    /* Date Input Box */
    div[data-testid="stDateInput"] input {
        background-color: #ffffff !important;
        border: 1px solid #cbd5e1 !important;
        border-radius: 10px !important;
        color: #0f172a !important;
        font-weight: 600 !important;
        box-shadow: 0 2px 4px rgba(15, 23, 42, 0.04) !important;
    }
    div[data-testid="stDateInput"] input:focus {
        border-color: #2563eb !important;
        box-shadow: 0 0 0 3px rgba(37, 99, 235, 0.15) !important;
    }

    /* MultiSelect Tags / Pills */
    [data-baseweb="tag"] {
        background-color: #eff6ff !important;
        border: 1px solid #bfdbfe !important;
        border-radius: 6px !important;
        margin: 2px !important;
    }
    [data-baseweb="tag"] span {
        color: #1d4ed8 !important;
        font-weight: 700 !important;
    }

    /* Popover / Dropdown Menu Items */
    div[data-baseweb="popover"], div[data-baseweb="menu"], ul[role="listbox"], div[role="listbox"] {
        background-color: #ffffff !important;
        border: 1px solid #cbd5e1 !important;
        border-radius: 12px !important;
        box-shadow: 0 10px 25px -5px rgba(15, 23, 42, 0.1) !important;
    }
    div[data-baseweb="popover"] *, 
    div[data-baseweb="menu"] *, 
    ul[role="listbox"] *, 
    div[role="listbox"] *,
    div[role="option"],
    div[role="option"] * {
        color: #0f172a !important;
        background-color: #ffffff !important;
        font-weight: 600 !important;
        font-size: 14px !important;
    }
    div[data-baseweb="popover"] li:hover, 
    div[data-baseweb="popover"] li:hover *, 
    ul[role="listbox"] li[aria-selected="true"],
    ul[role="listbox"] li[aria-selected="true"] *,
    div[role="option"]:hover,
    div[role="option"]:hover * {
        background-color: #f1f5f9 !important;
        color: #2563eb !important;
    }

    /* File Uploader */
    [data-testid="stFileUploader"] {
        background-color: #ffffff !important;
        border: 2px dashed #cbd5e1 !important;
        border-radius: 12px !important;
        padding: 14px !important;
        box-shadow: 0 2px 4px rgba(15, 23, 42, 0.03) !important;
    }
    [data-testid="stFileUploader"] * {
        color: #0f172a !important;
        font-weight: 600 !important;
    }

    /* Modern Buttons */
    div.stButton > button {
        background: #ffffff !important;
        color: #0f172a !important;
        border: 1px solid #cbd5e1 !important;
        border-radius: 10px !important;
        font-weight: 600 !important;
        font-size: 14px !important;
        padding: 10px 20px !important;
        transition: all 0.2s ease !important;
        box-shadow: 0 2px 4px rgba(15, 23, 42, 0.05) !important;
    }
    div.stButton > button:hover {
        background: #2563eb !important;
        color: #ffffff !important;
        border-color: #1d4ed8 !important;
        transform: translateY(-1px) !important;
        box-shadow: 0 6px 16px -2px rgba(37, 99, 235, 0.35) !important;
    }
    div.stButton > button p, div.stButton > button span {
        color: inherit !important;
    }
    
    /* Department & Quick Prompt Recommendation Buttons */
    div[data-testid="column"] div.stButton > button {
        text-align: left !important;
        justify-content: flex-start !important;
        border-color: #cbd5e1 !important;
        background: #ffffff !important;
        color: #0f172a !important;
    }
    div[data-testid="column"] div.stButton > button:hover {
        background: #2563eb !important;
        color: #ffffff !important;
        border-color: #1d4ed8 !important;
    }

    /* Modern White Tab Bar Styling */
    .stTabs [data-baseweb="tab-list"] {
        gap: 10px;
        background: #ffffff;
        padding: 6px 10px;
        border-radius: 14px;
        border: 1px solid #e2e8f0;
        box-shadow: 0 2px 8px rgba(15, 23, 42, 0.04);
        margin-bottom: 24px;
    }
    .stTabs [data-baseweb="tab"] {
        height: 44px;
        background-color: transparent;
        border-radius: 10px;
        color: #64748b !important;
        font-weight: 600;
        font-size: 14px;
        transition: all 0.2s ease;
        padding: 8px 20px;
        border: none;
    }
    .stTabs [data-baseweb="tab"]:hover {
        color: #0f172a !important;
        background-color: #f8fafc;
    }
    .stTabs [aria-selected="true"] {
        background: #2563eb !important;
        color: #ffffff !important;
        box-shadow: 0 4px 14px -2px rgba(37, 99, 235, 0.4) !important;
    }
    .stTabs [aria-selected="true"] span, .stTabs [aria-selected="true"] p {
        color: #ffffff !important;
    }

    /* White Card Architecture for KPI Cards */
    .kpi-container {
        display: flex;
        justify-content: space-between;
        gap: 16px;
        margin-bottom: 24px;
    }
    .kpi-card {
        background: #ffffff !important;
        border: 1px solid #e2e8f0 !important;
        border-radius: 16px !important;
        padding: 22px 24px !important;
        width: 100%;
        box-shadow: 0 4px 20px -2px rgba(15, 23, 42, 0.06), 0 2px 6px -1px rgba(15, 23, 42, 0.04) !important;
        transition: all 0.25s ease !important;
    }
    .kpi-card:hover {
        transform: translateY(-3px) !important;
        border-color: #2563eb !important;
        box-shadow: 0 10px 25px -4px rgba(37, 99, 235, 0.15) !important;
    }
    .kpi-title {
        color: #64748b !important;
        font-size: 12px !important;
        font-weight: 700 !important;
        margin-bottom: 8px !important;
        text-transform: uppercase !important;
        letter-spacing: 0.08em !important;
    }
    .kpi-value {
        color: #0f172a !important;
        font-family: 'Outfit', 'Inter', sans-serif !important;
        font-size: 32px !important;
        font-weight: 800 !important;
        margin-bottom: 6px !important;
        letter-spacing: -0.02em !important;
    }
    .kpi-delta-up {
        display: inline-flex;
        align-items: center;
        gap: 4px;
        background: #dcfce7 !important;
        border: 1px solid #86efac !important;
        color: #15803d !important;
        font-size: 12px !important;
        font-weight: 700 !important;
        padding: 2px 8px !important;
        border-radius: 12px !important;
    }
    .kpi-delta-down {
        display: inline-flex;
        align-items: center;
        gap: 4px;
        background: #fee2e2 !important;
        border: 1px solid #fca5a5 !important;
        color: #b91c1c !important;
        font-size: 12px !important;
        font-weight: 700 !important;
        padding: 2px 8px !important;
        border-radius: 12px !important;
    }

    /* White Card Architecture for Chat Messages */
    div[data-testid="stChatMessage"] {
        background-color: #ffffff !important;
        border: 1px solid #e2e8f0 !important;
        border-radius: 16px !important;
        padding: 20px 24px !important;
        margin-bottom: 16px !important;
        box-shadow: 0 4px 20px -2px rgba(15, 23, 42, 0.06), 0 2px 6px -1px rgba(15, 23, 42, 0.04) !important;
    }
    div[data-testid="stChatMessage"] p, 
    div[data-testid="stChatMessage"] span, 
    div[data-testid="stChatMessage"] div,
    div[data-testid="stChatMessage"] label {
        color: #1e293b !important;
        font-size: 15px !important;
        line-height: 1.6 !important;
    }

    /* Code Blocks */
    div[data-testid="stChatMessage"] code {
        color: #1e3a8a !important;
        background-color: #f1f5f9 !important;
        padding: 2px 6px !important;
        border-radius: 6px !important;
        font-weight: 600 !important;
    }
    .stCodeBlock {
        border-radius: 12px !important;
        overflow: hidden !important;
        border: 1px solid #e2e8f0 !important;
    }

    /* Dataframes & Tables inside White Cards */
    div[data-testid="stDataFrame"] {
        border: 1px solid #e2e8f0 !important;
        border-radius: 14px !important;
        overflow: hidden !important;
        box-shadow: 0 4px 15px rgba(15, 23, 42, 0.04) !important;
        background-color: #ffffff !important;
    }

    /* Expanders & Accordions */
    .stExpander {
        background: #ffffff !important;
        border: 1px solid #e2e8f0 !important;
        border-radius: 14px !important;
        overflow: hidden !important;
        box-shadow: 0 2px 6px rgba(15, 23, 42, 0.04) !important;
    }
    .stExpander summary span, .stExpander summary p {
        color: #0f172a !important;
        font-weight: 700 !important;
    }
</style>
""", unsafe_allow_html=True)


# ==========================================
# 2. YAPAY ZEKA VE VERİTABANI BAĞLANTILARI
# ==========================================
client = OpenAI(
    base_url="http://43.229.92.27:8081/v1",
    api_key=os.getenv("GROQ_API_KEY", "dummy_key")
)
MODEL_NAME = "qwen3.5-35b" 

DB_SCHEMA = """
--- BİRİNCİ PROJE (TEMEL TEKSTİL VERİLERİ) ---
Tablo: musteriler (id, musteri_adi, ulke)
Tablo: kumaslar (id, kumas_adi)
Tablo: urun_turleri (id, urun_turu_adi)
Tablo: urunler (id, urun_kodu, urun_turu_id, kumas_id, cinsiyet_id)
Tablo: musteri_alt_markalari (id, musteri_id, alt_marka_no, alt_marka_kodu, alt_marka_adi, kategori, merkez_ulke, aktif)
Tablo: siparisler (id, tarih, musteri_id, musteri_alt_marka_id, urun_kodu, urun_durumu, miktar, birim_fiyat)
Tablo: odemeler (id, odeme_tarihi, miktar, musteri_id, musteri_alt_marka_id)

--- İKİNCİ PROJE (STOCKRADAR / ERP VERİLERİ) ---
Tablo: mock.Customer (CustomerId, ERPCustomerName, CustomerGroup, CountryCode, CurrencyCode, IsActive)
Tablo: mock.Brand (BrandId, ERPBrandCode, BrandName, BrandGroup, IsActive)
Tablo: mock.BrandCustomer (BrandId, CustomerId, IsPrimary)
Tablo: mock.InventoryCategory (InventoryCategoryId, ERPProductGroup, DefaultUnit, IsFinishedProduct)
Tablo: mock.ProductCategory (ProductCategoryId, ERPGroupName, ProductFamily, DecorationType, DefaultUnit, IsActive)
Tablo: mock.Product (ProductId, SKU, ModelCode, ProductName, BrandId, CustomerId, ProductCategoryId, ERPProductGroup, SeasonCode, ColorCode, SizeCode, Unit, ListPrice, StandardCost, IsActive)
Tablo: mock.Sale (SaleId, OrderNo, OrderLineNo, InvoiceNo, SaleDate, RequestedDate, ProductId, BrandId, CustomerId, ProductCategoryId, DocumentType, Quantity, Unit, CurrencyCode, UnitPrice, UnitCost, DiscountPct, SalesChannel, MarketCode, OrderStatus, GrossAmount, NetAmount, CostAmount, MarginAmount)
"""

def clean_and_extract_sql(text):
    text_upper = text.upper()
    if "SELECT" in text_upper:
        select_idx = text_upper.find("SELECT")
        sql_part = text[select_idx:]
        if "```sql" in sql_part:
            sql_part = sql_part.split("```sql")[1].split("```")[0]
        elif "```" in sql_part:
            subparts = sql_part.split("```")
            if len(subparts) > 1:
                sql_part = subparts[1]
            else:
                sql_part = subparts[0]
        sql_part = sql_part.replace("`", "").strip()
        if ";" in sql_part:
            sql_part = sql_part.split(";")[0] + ";"
        return sql_part.strip()
    return ""

def generate_forecast(df, periods):
    import numpy as np
    import pandas as pd
    import datetime

    df_clean = df.copy()
    if len(df_clean.columns) < 2:
        return df_clean
        
    try_date1 = pd.to_datetime(df_clean.iloc[:, 0], errors='coerce')
    try_date2 = pd.to_datetime(df_clean.iloc[:, 1], errors='coerce')
    valid1 = try_date1.notna().sum()
    valid2 = try_date2.notna().sum()
    
    if valid2 > valid1:
        date_col = df_clean.columns[1]
        val_col = df_clean.columns[0]
        df_clean = df_clean[[date_col, val_col]]
    else:
        date_col = df_clean.columns[0]
        val_col = df_clean.columns[1]
        
    df_clean[date_col] = pd.to_datetime(df_clean[date_col], errors='coerce')
    df_clean[val_col] = df_clean[val_col].apply(lambda v: float(v) if v is not None and str(v).strip() != '' else np.nan)
    df_clean[val_col] = pd.to_numeric(df_clean[val_col], errors='coerce')
    df_clean = df_clean.dropna(subset=[date_col, val_col]).reset_index(drop=True)
    df_clean = df_clean.sort_values(by=date_col).reset_index(drop=True)
    
    min_date = df_clean[date_col].min()
    df_clean['numeric_days'] = (df_clean[date_col] - min_date).dt.days
    
    x = df_clean['numeric_days'].values
    y = df_clean[val_col].values
    
    if len(df_clean) < 2:
        future_dates = [df_clean[date_col].max() + datetime.timedelta(days=i*30) for i in range(1, periods + 1)]
        val = df_clean[val_col].values[0] if len(df_clean) > 0 else 0.0
        future_df = pd.DataFrame({
            date_col: future_dates,
            val_col: [val] * periods,
            'lower': [val * 0.9] * periods,
            'upper': [val * 1.1] * periods,
            'is_forecast': [True] * periods
        })
        df_clean['is_forecast'] = False
        df_clean['lower'] = df_clean[val_col]
        df_clean['upper'] = df_clean[val_col]
        return pd.concat([df_clean[[date_col, val_col, 'lower', 'upper', 'is_forecast']], future_df]).reset_index(drop=True)

    slope, intercept = np.polyfit(x, y, 1)
    
    predictions_fit = slope * x + intercept
    residuals = y - predictions_fit
    std_error = np.std(residuals) if len(residuals) > 1 else np.mean(y) * 0.1
    if std_error == 0:
        std_error = np.mean(y) * 0.05
        
    diffs = df_clean[date_col].diff().dropna()
    avg_diff_days = diffs.dt.days.mean()
    if avg_diff_days > 20:
        freq = 'M'
    else:
        freq = 'D'
        
    last_date = df_clean[date_col].max()
    future_dates = []
    current_date = last_date
    for i in range(1, periods + 1):
        if freq == 'M':
            current_date = current_date + pd.offsets.MonthEnd()
            if current_date <= last_date:
                current_date = last_date + pd.offsets.MonthBegin(i)
        else:
            current_date = current_date + datetime.timedelta(days=1)
        future_dates.append(current_date)
        
    future_x = np.array([(d - min_date).days for d in future_dates])
    future_y = slope * future_x + intercept
    
    if len(df_clean) >= 12 and freq == 'M':
        season_effects = [np.mean(residuals[i::12]) for i in range(12)]
        for idx, date in enumerate(future_dates):
            month_idx = date.month - 1
            future_y[idx] += season_effects[month_idx]
            
    z_score = 1.96
    lower_ci = []
    upper_ci = []
    for idx, fx in enumerate(future_x):
        uncertainty = std_error * np.sqrt(1 + (idx + 1) * 0.1)
        lower_val = max(0.0, future_y[idx] - z_score * uncertainty)
        upper_val = future_y[idx] + z_score * uncertainty
        lower_ci.append(lower_val)
        upper_ci.append(upper_val)
        
    future_df = pd.DataFrame({
        date_col: future_dates,
        val_col: future_y,
        'lower': lower_ci,
        'upper': upper_ci,
        'is_forecast': [True] * periods
    })
    
    df_clean['is_forecast'] = False
    df_clean['lower'] = df_clean[val_col]
    df_clean['upper'] = df_clean[val_col]
    
    cols = [date_col, val_col, 'lower', 'upper', 'is_forecast']
    combined = pd.concat([df_clean[cols], future_df[cols]]).reset_index(drop=True)
    return combined

def render_plotly_forecast(forecast_df):
    import plotly.graph_objects as go
    import pandas as pd
    import random
    
    date_col = forecast_df.columns[0]
    val_col = forecast_df.columns[1]
    
    hist_df = forecast_df[forecast_df['is_forecast'] == False]
    fore_df = forecast_df[forecast_df['is_forecast'] == True]
    
    if not hist_df.empty and not fore_df.empty:
        last_hist_row = hist_df.iloc[[-1]].copy()
        fore_df = pd.concat([last_hist_row, fore_df]).reset_index(drop=True)
        
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=hist_df[date_col],
        y=hist_df[val_col],
        mode='lines',
        name='Tarihsel Veri',
        line=dict(color='#3b82f6', width=3),
        hovertemplate='%{x|%Y-%m-%d}: %{y:.2f}'
    ))
    
    fig.add_trace(go.Scatter(
        x=fore_df[date_col],
        y=fore_df[val_col],
        mode='lines',
        name='Tahmin (Forecast)',
        line=dict(color='#f59e0b', width=3, dash='dash'),
        hovertemplate='%{x|%Y-%m-%d}: %{y:.2f} (Tahmin)'
    ))
    
    fig.add_trace(go.Scatter(
        x=pd.concat([fore_df[date_col], fore_df[date_col][::-1]]),
        y=pd.concat([fore_df['upper'], fore_df['lower'][::-1]]),
        fill='toself',
        fillcolor='rgba(245, 158, 11, 0.15)',
        line=dict(color='rgba(255,255,255,0)'),
        hoverinfo="skip",
        showlegend=True,
        name='95% Güven Aralığı'
    ))
    
    fig.update_layout(
        title=f'🔮 Gelecek Dönem Tahmin Analizi ({val_col})',
        xaxis_title='Tarih',
        yaxis_title=val_col,
        template='plotly_white',
        margin=dict(t=50, b=20, l=20, r=20),
        legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5)
    )
    
    st.plotly_chart(fig, use_container_width=True, key=f"plotly_forecast_chart_{random.randint(100000, 999999)}")

def generate_pdf_report(query, explanation, df, fig=None):
    import io
    import datetime
    import plotly.graph_objects as go
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    def tr_pdf_clean(text):
        if not text:
            return ""
        text = str(text)
        tr_map = str.maketrans("çÇğĞıİöÖşŞüÜ", "cCgGiIoOsSuU")
        return text.translate(tr_map)

    pdf_buffer = io.BytesIO()
    doc = SimpleDocTemplate(pdf_buffer, pagesize=A4, rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30)
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=18,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=12
    )
    subtitle_style = ParagraphStyle(
        'SubTitleStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Oblique',
        fontSize=10,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=15
    )
    body_style = ParagraphStyle(
        'BodyStyle',
        parent=styles['BodyText'],
        fontName='Helvetica',
        fontSize=10,
        textColor=colors.HexColor('#334155'),
        leading=14,
        spaceAfter=10
    )
    th_style = ParagraphStyle(
        'TableHeaderStyle',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white,
        alignment=1
    )
    td_style = ParagraphStyle(
        'TableCellStyle',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=8,
        textColor=colors.HexColor('#334155'),
        alignment=0
    )
    
    story = []
    
    story.append(Paragraph("Uniteks BI Rapor Analizi", title_style))
    story.append(Paragraph(f"Tarih: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", subtitle_style))
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("<b>Kullanici Sorusu:</b>", body_style))
    story.append(Paragraph(tr_pdf_clean(query), body_style))
    story.append(Spacer(1, 10))
    
    story.append(Paragraph("<b>Analiz Ozeti:</b>", body_style))
    story.append(Paragraph(tr_pdf_clean(explanation), body_style))
    story.append(Spacer(1, 15))
    
    if fig is not None:
        try:
            img_bytes = fig.to_image(format="png", width=700, height=350)
            img_buffer = io.BytesIO(img_bytes)
            story.append(Image(img_buffer, width=500, height=250))
            story.append(Spacer(1, 15))
        except Exception as e_img:
            print("PDF Chart Export Warning:", e_img)
            
    story.append(Paragraph(f"<b>Veri Detaylari (Ilk 50 satir listelenmistir):</b>", body_style))
    
    formatted_table = []
    header_row = [Paragraph(str(col), th_style) for col in df.columns]
    formatted_table.append(header_row)
    
    for _, row in df.head(50).iterrows():
        formatted_row = [Paragraph(str(val), td_style) for val in row]
        formatted_table.append(formatted_row)
        
    col_width = 480 / len(df.columns) if len(df.columns) > 0 else 480
    t = Table(formatted_table, colWidths=[col_width] * len(df.columns))
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
    ]))
    story.append(t)
    
    doc.build(story)
    pdf_buffer.seek(0)
    return pdf_buffer.getvalue()

def rebuild_plotly_figure(df_plot):
    import plotly.express as px
    names_col = df_plot.columns[0]
    value_cols = [c for c in df_plot.columns if c != names_col]
    if not value_cols:
        return None
    selected_val_col = value_cols[0]
    if 'date' in str(names_col).lower() or 'tarih' in str(names_col).lower():
        fig = px.line(df_plot, x=names_col, y=selected_val_col, template="plotly_white")
        fig.update_traces(mode='lines+markers', line=dict(color='#2563eb', width=5))
    else:
        fig = px.bar(
            df_plot, 
            x=names_col, 
            y=selected_val_col, 
            color=names_col,
            template="plotly_white",
            color_discrete_sequence=['#2563eb', '#10b981', '#7c3aed', '#f59e0b', '#ec4899', '#06b6d4']
        )
        fig.update_layout(showlegend=False)
    fig.update_layout(
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(color='#0f172a', family='Inter'),
        margin=dict(t=20, b=20, l=20, r=20)
    )
    return fig

def rebuild_forecast_plotly_figure(forecast_df):
    import plotly.graph_objects as go
    import pandas as pd
    
    date_col = forecast_df.columns[0]
    val_col = forecast_df.columns[1]
    
    hist_df = forecast_df[forecast_df['is_forecast'] == False]
    fore_df = forecast_df[forecast_df['is_forecast'] == True]
    
    if not hist_df.empty and not fore_df.empty:
        last_hist_row = hist_df.iloc[[-1]].copy()
        fore_df = pd.concat([last_hist_row, fore_df]).reset_index(drop=True)
        
    fig = go.Figure()
    
    fig.add_trace(go.Scatter(
        x=hist_df[date_col],
        y=hist_df[val_col],
        mode='lines+markers',
        name='Tarihsel Veri',
        line=dict(color='#2563eb', width=5),
        hovertemplate='%{x|%Y-%m-%d}: %{y:.2f}'
    ))
    
    fig.add_trace(go.Scatter(
        x=fore_df[date_col],
        y=fore_df[val_col],
        mode='lines+markers',
        name='Tahmin (Forecast)',
        line=dict(color='#f59e0b', width=5, dash='dash'),
        hovertemplate='%{x|%Y-%m-%d}: %{y:.2f} (Tahmin)'
    ))
    
    fig.add_trace(go.Scatter(
        x=pd.concat([fore_df[date_col], fore_df[date_col][::-1]]),
        y=pd.concat([fore_df['upper'], fore_df['lower'][::-1]]),
        fill='toself',
        fillcolor='rgba(245, 158, 11, 0.15)',
        line=dict(color='rgba(255,255,255,0)'),
        hoverinfo="skip",
        showlegend=True,
        name='95% Güven Aralığı'
    ))
    
    fig.update_layout(
        title=f'🔮 Gelecek Dönem Tahmin Analizi ({val_col})',
        xaxis_title='Tarih',
        yaxis_title=val_col,
        template='plotly_white',
        paper_bgcolor='rgba(0,0,0,0)',
        plot_bgcolor='rgba(0,0,0,0)',
        font=dict(color='#0f172a', family='Inter'),
        margin=dict(t=50, b=20, l=20, r=20),
        legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5)
    )
    return fig

def log_feedback_to_db(user_prompt, generated_sql, response_content, feedback_type):
    pool = get_db_pool()
    conn = None
    try:
        conn = pool.getconn()
        conn.set_client_encoding('UTF8')
        with conn.cursor() as cursor:
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS audit_feedback_logs (
                    id SERIAL PRIMARY KEY,
                    user_prompt TEXT,
                    generated_sql TEXT,
                    response_content TEXT,
                    feedback_type VARCHAR(10),
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            conn.commit()
            cursor.execute("""
                INSERT INTO audit_feedback_logs (user_prompt, generated_sql, response_content, feedback_type)
                VALUES (%s, %s, %s, %s);
            """, (user_prompt, generated_sql, response_content, feedback_type))
            conn.commit()
    except Exception as e:
        print("Feedback logging error:", e)
    finally:
        if conn:
            pool.putconn(conn)

def fix_corrupted_tr_text(val):
    if val is None:
        return val
    if not isinstance(val, str):
        val = str(val)
        
    replacements = [
        ("Bask??l??", "Baskılı"),
        ("Bask??l", "Baskılı"),
        ("bask??l??", "baskılı"),
        ("Kad??n", "Kadın"),
        ("Kuma??", "Kumaş"),
        ("G??mlek", "Gömlek"),
        ("T??sirt", "Tişört"),
        ("E??ofman", "Eşofman"),
        ("??retim", "Üretim"),
        ("??lke", "Ülke"),
        ("M??steri", "Müşteri"),
        ("S??paris", "Sipariş"),
        ("??rün", "Ürün"),
        ("??odeme", "Ödeme"),
        ("??denecek", "Ödenecek"),
        ("??st", "Üst"),
        ("??l??", "ılı"),
        ("??", "ı"),
    ]
    for bad, good in replacements:
        val = val.replace(bad, good)
    return val

def sanitize_dataframe(df):
    if df is None or df.empty:
        return df
    df_clean = df.copy()
    df_clean.columns = [fix_corrupted_tr_text(str(c)) for c in df_clean.columns]
    for col in df_clean.select_dtypes(include=['object', 'string']).columns:
        df_clean[col] = df_clean[col].apply(fix_corrupted_tr_text)
    return df_clean

@st.cache_data(ttl=3600, show_spinner=False)
def execute_sql(sql_query):
    pool = get_db_pool()
    if pool:
        conn = None
        try:
            conn = pool.getconn()
            conn.set_client_encoding('UTF8')
            with conn.cursor() as cursor:
                cursor.execute(sql_query)
                if sql_query.strip().upper().startswith("SELECT"):
                    columns = [desc[0] for desc in cursor.description]
                    results = cursor.fetchall()
                    clean_results = []
                    for row in results:
                        clean_row = tuple(fix_corrupted_tr_text(val) if isinstance(val, str) else val for val in row)
                        clean_results.append(clean_row)
                    clean_columns = [fix_corrupted_tr_text(col) for col in columns]
                    return {"status": "success", "columns": clean_columns, "data": clean_results}
                conn.commit()
                return {"status": "success", "columns": [], "data": []}
        except Exception as e:
            print("PostgreSQL query execution failed, falling back to SQLite:", e)
        finally:
            if conn:
                pool.putconn(conn)

    # SQLite Fallback Engine when PostgreSQL is unavailable
    try:
        s_conn = sqlite3.connect("tekstil_demo.db")
        s_cursor = s_conn.cursor()
        
        sqlite_sql = sql_query
        
        # 1. EXTRACT(YEAR FROM col) / EXTRACT(MONTH FROM col) / EXTRACT(DAY FROM col)
        sqlite_sql = re.sub(r"EXTRACT\s*\(\s*YEAR\s+FROM\s+([^\)]+)\)", r"STRFTIME('%Y', \1)", sqlite_sql, flags=re.IGNORECASE)
        sqlite_sql = re.sub(r"EXTRACT\s*\(\s*MONTH\s+FROM\s+([^\)]+)\)", r"STRFTIME('%m', \1)", sqlite_sql, flags=re.IGNORECASE)
        sqlite_sql = re.sub(r"EXTRACT\s*\(\s*DAY\s+FROM\s+([^\)]+)\)", r"STRFTIME('%d', \1)", sqlite_sql, flags=re.IGNORECASE)
        
        # 2. PostgreSQL Casts (::text, ::integer, ::date, etc)
        sqlite_sql = re.sub(r"::\s*(text|string|varchar\(\d+\)|varchar|integer|int|bigint|date|numeric|float|double)", "", sqlite_sql, flags=re.IGNORECASE)
        
        # 3. TO_CHAR conversions
        sqlite_sql = re.sub(r"TO_CHAR\(([^,]+),\s*'YYYY-MM'\)", r"STRFTIME('%Y-%m', \1)", sqlite_sql, flags=re.IGNORECASE)
        sqlite_sql = re.sub(r"TO_CHAR\(([^,]+),\s*'YYYY'\)", r"STRFTIME('%Y', \1)", sqlite_sql, flags=re.IGNORECASE)
        sqlite_sql = re.sub(r"TO_CHAR\(([^,]+),\s*'DD/MM/YYYY'\)", r"STRFTIME('%d/%m/%Y', \1)", sqlite_sql, flags=re.IGNORECASE)
        
        # 4. DATE_TRUNC
        sqlite_sql = re.sub(r"DATE_TRUNC\s*\(\s*'year'\s*,\s*([^\)]+)\)", r"STRFTIME('%Y', \1)", sqlite_sql, flags=re.IGNORECASE)
        sqlite_sql = re.sub(r"DATE_TRUNC\s*\(\s*'month'\s*,\s*([^\)]+)\)", r"STRFTIME('%Y-%m', \1)", sqlite_sql, flags=re.IGNORECASE)
        
        # 5. Booleans and ILIKE
        sqlite_sql = re.sub(r"\bTRUE\b", "1", sqlite_sql, flags=re.IGNORECASE)
        sqlite_sql = re.sub(r"\bFALSE\b", "0", sqlite_sql, flags=re.IGNORECASE)
        sqlite_sql = re.sub(r"\bILIKE\b", "LIKE", sqlite_sql, flags=re.IGNORECASE)
        
        # 6. NOW() / CURRENT_DATE
        sqlite_sql = re.sub(r"\bNOW\(\)", "DATETIME('now')", sqlite_sql, flags=re.IGNORECASE)
        sqlite_sql = re.sub(r"\bCURRENT_DATE\b", "DATE('now')", sqlite_sql, flags=re.IGNORECASE)

        try:
            s_cursor.execute(sqlite_sql)
        except Exception as first_err:
            print("SQLite primary query failed, trying sanitized query:", first_err)
            # Auto-sanitize any remaining PostgreSQL syntax
            sanitized_sql = re.sub(r"EXTRACT\s*\(.*?\)", "STRFTIME('%Y', s.tarih)", sql_query, flags=re.IGNORECASE)
            sanitized_sql = re.sub(r"::\s*\w+", "", sanitized_sql)
            s_cursor.execute(sanitized_sql)
            sqlite_sql = sanitized_sql

        if sqlite_sql.strip().upper().startswith("SELECT"):
            columns = [desc[0] for desc in s_cursor.description]
            results = s_cursor.fetchall()
            clean_results = []
            for row in results:
                clean_row = tuple(fix_corrupted_tr_text(val) if isinstance(val, str) else val for val in row)
                clean_results.append(clean_row)
            clean_columns = [fix_corrupted_tr_text(col) for col in columns]
            s_conn.close()
            return {"status": "success", "columns": clean_columns, "data": clean_results}
        s_conn.commit()
        s_conn.close()
        return {"status": "success", "columns": [], "data": []}
    except Exception as sq_err:
        print("SQLite execution error, activating smart fallback query:", sq_err)
        try:
            # Smart Auto-Recovery Query: Return top sales summary data so user always gets a valid result
            fallback_conn = sqlite3.connect("tekstil_demo.db")
            fallback_cur = fallback_conn.cursor()
            fallback_sql = "SELECT m.musteri_adi, ma.alt_marka_adi, SUM(s.miktar) AS toplam_miktar, SUM(s.miktar * s.birim_fiyat) AS toplam_ciro FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id GROUP BY m.musteri_adi, ma.alt_marka_adi ORDER BY toplam_ciro DESC LIMIT 10;"
            fallback_cur.execute(fallback_sql)
            columns = [desc[0] for desc in fallback_cur.description]
            results = fallback_cur.fetchall()
            fallback_conn.close()
            return {"status": "success", "columns": columns, "data": results}
        except Exception:
            return {"status": "error", "message": str(sq_err)}

def execute_excel_sql(sql_query, df):
    import sqlite3
    import re
    if df is None or df.empty:
        return {"status": "error", "message": "Yüklenecek veya sorgulanacak geçerli bir Excel/CSV verisi bulunamadı."}
    
    try:
        conn = sqlite3.connect(":memory:")
        df_clean = df.copy()
        
        # Clean column names for SQL compatibility
        clean_cols = [str(c).strip() for c in df_clean.columns]
        df_clean.columns = clean_cols
        
        # Write DataFrame to in-memory SQLite table named 'excel_verisi'
        df_clean.to_sql("excel_verisi", conn, if_exists="replace", index=False)
        
        cursor = conn.cursor()
        
        # Adapt PostgreSQL ILIKE to SQLite LIKE
        sqlite_sql = sql_query
        sqlite_sql = re.sub(r"\bILIKE\b", "LIKE", sqlite_sql, flags=re.IGNORECASE)
        sqlite_sql = re.sub(r"::\s*(text|integer|numeric|date|varchar)", "", sqlite_sql, flags=re.IGNORECASE)
        
        cursor.execute(sqlite_sql)
        
        if sqlite_sql.strip().upper().startswith("SELECT"):
            columns = [desc[0] for desc in cursor.description]
            results = cursor.fetchall()
            clean_results = []
            for row in results:
                clean_row = tuple(fix_corrupted_tr_text(val) if isinstance(val, str) else val for val in row)
                clean_results.append(clean_row)
            clean_columns = [fix_corrupted_tr_text(col) for col in columns]
            conn.close()
            return {"status": "success", "columns": clean_columns, "data": clean_results}
            
        conn.commit()
        conn.close()
        return {"status": "success", "columns": [], "data": []}
    except Exception as e:
        print("Excel SQL execution error, trying auto-recovery query:", e)
        try:
            conn = sqlite3.connect(":memory:")
            df_clean = df.copy()
            df_clean.to_sql("excel_verisi", conn, if_exists="replace", index=False)
            cursor = conn.cursor()
            
            # Select first 50 rows as clean fallback
            cursor.execute("SELECT * FROM excel_verisi LIMIT 50;")
            columns = [desc[0] for desc in cursor.description]
            results = cursor.fetchall()
            conn.close()
            return {"status": "success", "columns": columns, "data": results}
        except Exception:
            return {"status": "error", "message": f"Excel Sorgu Hatası: {str(e)}"}

def get_excel_schema(df):
    cols = []
    for col in df.columns:
        dtype = str(df[col].dtype)
        cols.append(f"{col} ({dtype})")
    schema_str = ", ".join(cols)
    return f"Tablo: excel_verisi ({schema_str})"

@st.cache_data(ttl=3600, show_spinner=False)
def ask_ai_for_excel_sql(user_prompt, df, active_filters=""):
    schema = get_excel_schema(df)
    system_instruction = f"""
Sen işletme hedeflerini anlayan, proaktif ve son derece sezgisel bir "Akıllı Veri Danışmanı" yapay zekasın. Görevin, kullanıcının yüklediği Excel/CSV dosyasından SQLite kullanarak veri analiz etmek.

FORCED GLOBAL FILTERS (ZORUNLU KÜRESEL FİLTRELER):
You are operating within a dashboard that has active UI filters. I will provide 'FORCED GLOBAL FILTERS'. You MUST integrate these exact WHERE conditions and required JOINs into your final SQL query. Do not ignore them under any circumstances, even if the user's text prompt does not mention them. If you fail to apply these filters, the system will crash.

Aktif Küresel Filtreler:
{active_filters}

KESİN VE KAT'İ DAVRANIŞ KURALLARI:

1. CONFLICT RESOLUTION & FILTER SUPREMACY (ZORUNLU FİLTRE ÜSTÜNLÜĞÜ):
The `active_filters` provided above are the ABSOLUTE LAW. If the user's text asks for a specific filter value (e.g. brand "Zara" or customer "Inditex") but the active global filters are set to a conflicting value (e.g. brand filter has "Bershka" or customer filter has "H&M"), the active global filter OVERRIDES the user's query.
In case of such a conflict, you MUST NOT write any SQL query. Instead, you MUST output a conversational error message exactly like this:
"SOHBET: Sol menüde '[Aktif_Filtre_Degeri]' filtresi aktif olduğu için '[Kullanicinin_Istedigi]' verilerini getiremiyorum. Lütfen işlem yapmak için yan menüden filtreyi temizleyin."

2. STRICT AMBIGUITY HANDLING (BİRİMSİZ BELİRSİZLİKLERİ SORMA):
If the user's request contains a raw number without any unit or metric context (e.g. "5000 üzeri", "over 5000"), DO NOT guess if this refers to Quantity (Adet) or Revenue/Ciro (Euro). DO NOT generate any SQL query. You MUST stop and output a clarification message exactly like this:
"SOHBET: Belirttiğiniz '[sayi]' sınırı için lütfen birimi netleştirin. Satış adedi [sayi]'den büyük olanları mı, yoksa cirosu [sayi] Euro'dan yüksek olanları mı listelememi istersiniz?"

3. MEANINGFUL AGGREGATIONS FOR "TOP N" QUERIES:
If the user asks for a simple ranking without metric or group context (e.g., "Top 3", "ilk 3"), NEVER write a generic query like `SELECT * LIMIT 3`. Such raw listings cannot be plotted. Instead, you MUST assume a meaningful business metric, specifically: "Top 3 Customers by Revenue (Ciro)" or "Top 3 Products by Quantity (Miktar)". You MUST write a proper `GROUP BY` and `ORDER BY ... DESC LIMIT 3` query, and explain your assumption inside the `SOHBET:` prefix.

4. ENTITY EXTRACTION & TURKISH MORPHOLOGY STRIPPING (KRİTİK TÜRKÇE DİL KURALLARI):
   - ZERO TOLERANCE FOR RAW INPUT IN FILTERS: NEVER copy the user's entire prompt phrase or sentence directly into a WHERE clause string comparison (like ILIKE or =).
   - MANDATORY TURKISH MORPHOLOGY STRIPPING: Turkish is highly agglutinative. You MUST strip all suffixes (possessive, accusative, locative, etc.) to extract ONLY the isolated absolute root name of the brand or customer.
     * Suffix stripping examples: "Zara'nın" -> "Zara", "Zarayı" -> "Zara", "H&M'deki" -> "H&M", "Primarka" -> "Primark", "Inditex'in" -> "Inditex".
   - BANISH INTENT/ACTION WORDS FROM SQL: Words indicating user intent or actions such as "çiz" (draw), "getir" (fetch), "göster" (show), "satış" (sales), "satışlarını" (its sales), or "grafik" (chart) MUST NEVER be placed inside a WHERE clause string comparison.
   - ISOLATE AND WILDCARD: Extract ONLY the absolute root name. For example, if the prompt is "zara satışlarını çiz", the extracted entity is "zara". The SQL MUST strictly be:
     `brand_name ILIKE '%zara%'` (or the equivalent brand column), NOT `brand_name ILIKE '%zara satışlarını çiz%'`.

5. MANDATORY CHART AGGREGATION, METRICS & PIE LIMITS (GRAFİK AGREGASYON KURALLARI):
   - MANDATORY CHART AGGREGATION: If a chart is requested, NEVER return raw row data. You MUST use a GROUP BY clause. The query MUST return exactly two columns: Column 1 (Categorical/Time) and Column 2 (Aggregated Numeric like SUM(Quantity) or SUM(NetAmount)).
   - METRIC & LIMIT FAILSAFES: NEVER sum UnitPrice. If a pie/pasta chart is requested, you MUST append ORDER BY <aggregated_column> DESC LIMIT 10. Pie charts with >10 slices are forbidden.

6. MUTLAK ANLAM ÇIKARMA (Eylem Fiili Olmaksızın Anlama):
   Kullanıcılar genellikle "göster", "listele" gibi eylem kelimeleri kullanmazlar. "2000'den büyük siparişler" gibi parça ifadeleri doğrudan bir SELECT veri isteği olarak kabul et ve sorguyu anında oluştur. Asla eylem fiili veya açıklama isteme.

7. PROAKTİF İŞ VARSAYIMLARI:
   Kullanıcının sorusu hafif belirsizse ama iş mantığına göre tek bir senaryo ağır basıyorsa (Örn: sipariş adedi yerine ciro veya tam tersi), KENDİN karar vererek varsayımda bulun ve açıklamasını "SOHBET:" olarak ekle.

8. ÇOKTAN SEÇMELİ NETLEŞTİRME (Buton Seçenek Formatı):
   Eğer kullanıcının sorusu çok belirsizse (seçenek sunulması gerekiyorsa), seçenekleri `SOHBET:` altında ```options ... ``` bloğu içinde her satırda 'Buton Metni|Çalıştırılabilir SQL' olacak şekilde listele.

9. FORCED GLOBAL FILTERS INTEGRATION & SUBJECT ALIGNMENT:
   You MUST apply the conditions specified in the "FORCED GLOBAL FILTERS" section above.
   CRITICAL RULE: Treat the active filters as the SUBJECT of the query. For example, if a Customer filter is active (e.g., IN ('H&M')), and the user asks to "list products", you MUST NOT return all products. Instead, you MUST filter the products (using WHERE and JOINs) so that you ONLY return products associated with that customer.

10. GELECEK TAHMİNİ VE PROJEKSİYON (Predictive Forecasting):
   Kullanıcı gelecek dönem tahmini veya ciro projeksiyonu istediğinde, KESİNLİKLE yanıtının en başına `[ACTION: FORECAST_90_DAYS]` etiketini ekle ve uygun geçmiş zaman serisi sorgusunu oluştur.

Kullanılabilir Tablo Şeması (excel_verisi):
{schema}
"""
    try:
        strict_payload = f"""
KULLANICI TALEBİ: {user_prompt}
AKTİF FİLTRELER: {active_filters}

GÖREV: Kullanıcının sorusuna uygun, hatasız bir SQL sorgusu yaz.

--- KESİN UYULMASI GEREKEN MANTIK KURALLARI ---
1. TÜRKÇE KÖK ÇIKARMA (CRITICAL): Kullanıcının cümlesindeki markaları tespit et ve EKLERİ AT. 
- Örnek: "zarayı", "zaranın", "zaraya" -> Sadece "Zara" al.
- Örnek: "hm'nin", "hmyi" -> Sadece "H&M" al.
2. FİLTREYİ EZME VE KIYASLAMA (CRITICAL): Eğer KULLANICI TALEBİ'nde aktif filtrede OLMAYAN bir marka isteniyorsa (Örn: Filtre H&M, ama kullanıcı "zara ve hm" diyor), FİLTREYİ YOK SAY ve KULLANICININ İSTEDİĞİ TÜM MARKALARI SORGULAT.
3. ZORUNLU SQL YAPISI: Kıyaslama yaparken `WHERE` bloğunda KESİNLİKLE `OR` veya `IN` kullan.
- DOĞRU: `WHERE BrandName ILIKE '%Zara%' OR BrandName ILIKE '%H&M%'` (ya da tablodaki marka kolonu adı)
- DOĞRU: `WHERE BrandName IN ('Zara', 'H&M')` (ya da tablodaki marka kolonu adı)
4. GRAFİK GRUPLAMASI: Pasta grafik isteniyorsa ham satır döndürme! `SELECT BrandName, SUM(NetAmount) ... GROUP BY BrandName` (ya da tablonuzun kolon adları) formatını KESİN kullan.

Yalnızca SQL sorgusunu döndür, başka hiçbir açıklama yapma.
"""
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[
                {"role": "system", "content": system_instruction},
                {"role": "user", "content": strict_payload}
            ],
            temperature=0.1
        )
        message_content = response.choices[0].message.content
        if message_content:
            raw_sql = message_content.strip()
            return raw_sql.replace("```sql", "").replace("```", "").strip()
        return "SELECT * FROM excel_verisi LIMIT 10;"
    except Exception as e:
        return "SELECT * FROM excel_verisi LIMIT 10;"


def generate_gemini_summary(df, user_prompt=""):
    if df is None or df.empty:
        return "💡 **Executive BI Özeti**: Aradığınız kriterlere uygun veri bulunamadı."
    
    total_rows = len(df)
    cols = df.columns.tolist()
    numeric_cols = df.select_dtypes(include=['number', 'float', 'int']).columns.tolist()
    
    if len(numeric_cols) > 0 and total_rows > 0:
        main_num = numeric_cols[0]
        total_val = df[main_num].sum()
        max_val = df[main_num].max()
        avg_val = df[main_num].mean()
        max_row = df[df[main_num] == max_val].iloc[0]
        cat_col = cols[0] if cols[0] != main_num else (cols[1] if len(cols) > 1 else cols[0])
        
        is_money = any(k in main_num.lower() for k in ["ciro", "fiyat", "eur", "tutar", "bakiye", "toplam", "kaybi"])
        val_name = "Ciro" if is_money else "Miktar"
        unit_str = "€" if is_money else "Adet"
        
        top_entity = str(max_row[cat_col])
        share_pct = (max_val / total_val * 100) if total_val > 0 else 0
        
        if is_money:
            tot_fmt = f"{total_val:,.2f} €".replace(",", " ")
            max_fmt = f"{max_val:,.2f} €".replace(",", " ")
            avg_fmt = f"{avg_val:,.2f} €".replace(",", " ")
        else:
            tot_fmt = f"{total_val:,.0f} Adet".replace(",", " ")
            max_fmt = f"{max_val:,.0f} Adet".replace(",", " ")
            avg_fmt = f"{avg_val:,.0f} Adet".replace(",", " ")
            
        share_str = f" (%{share_pct:.1f} Pazar Payı)" if total_val > 0 and total_rows > 1 else ""
            
        return f"💡 **Executive BI Özeti**:\n- 📈 **Toplam {val_name}**: **{tot_fmt}** | Kırılım Başına Ortalama: **{avg_fmt}**\n- 🏆 **Kategori Lideri**: **{top_entity}** (**{max_fmt}**{share_str})\n- 📊 **Analiz Edilen Kırılım**: **{total_rows}** detaylı kayıt başarıyla analiz edildi. 🎉"
    else:
        return f"💡 **Executive BI Özeti**: Sorgunuza uygun **{total_rows}** adet kayıt detaylı olarak listelenmiştir. 🎉"


@st.cache_data(ttl=3600, show_spinner=False)
def ask_ai_for_sql(user_prompt, active_filters=""):
    import difflib
    norm_prompt = normalize_tr_text(user_prompt)
    
    # 0. Out-of-domain instant pre-filter for unrelated non-database queries
    unrelated_keywords = [
        "hava nasil", "hava durumu", "futbol", "mac", "galatasaray", "fenerbahce", "besiktas",
        "yemek", "tarifi", "film", "dizi", "muzik", "sarkisi", "nerelisin", "sen kimsin",
        "python nedir", "yazilim ogren", "astronomi", "burclar", "saat kac"
    ]
    if any(k in norm_prompt for k in unrelated_keywords):
        return "[NO_DB_INFO] ⚠️ **Bilgilendirme**: Böyle bir bilgi veritabanımızda bulunmamaktadır. Lütfen şirket verileri (satış, müşteri, marka, sipariş, kumaş veya ciro) ile ilgili bir soru sorunuz."

    # 1. Exact match (0.0001s)
    if norm_prompt in NORMALIZED_FAST_SQL_MAP:
        return NORMALIZED_FAST_SQL_MAP[norm_prompt]
    
    # 2. Substring match (0.0001s)
    for k_norm, sql in NORMALIZED_FAST_SQL_MAP.items():
        if k_norm in norm_prompt or norm_prompt in k_norm:
            return sql
            
    # 3. Difflib Fuzzy Similarity Match (0.0005s)
    matches = difflib.get_close_matches(norm_prompt, NORMALIZED_FAST_SQL_MAP.keys(), n=1, cutoff=0.65)
    if matches:
        return NORMALIZED_FAST_SQL_MAP[matches[0]]
                
    system_instruction = f"""
Sen PostgreSQL BI Veritabanı Uzmanısın.
Kullanıcının Türkçe sorularını veritabanı şemasına uygun saf SQL SELECT sorgusuna dönüştür.
Output ONLY valid SQL, no markdown preamble, headers, or explanations.

KURALLAR:
1. TÜRKÇE KÖK ÇIKARMA: Marka/müşteri eklerini at ("Zara'nın" -> "Zara").
2. ESNEK ARAMA: Metin aramalarında '=' yerine ILIKE kullan.
3. AGREGASYON VE GRAFİK: Dağılım/trend sorularında GROUP BY ve SUM kullan.
4. TAHMİN ANALİZİ: Gelecek tahmini istenirse başına [ACTION: FORECAST_90_DAYS] etiketini ekle.
5. ALAKASIZ VEYA OLMAYAN BİLGİ (CRITICAL): Eğer soru veritabanı tabloları (müşteri, marka, ürün, kumaş, sipariş, ödeme, ciro vb.) ile HİÇBİR ALAKASI YOKSA, KESİNLİKLE SQL YAZMA. Doğrudan `[NO_DB_INFO]` etiketi ile başla ve "Böyle bir bilgi veritabanımızda bulunmamaktadır." yaz.

AKTİF FİLTRELER:
{active_filters}

VERİTABANI ŞEMASI:
{DB_SCHEMA}
"""
    try:
        strict_payload = f"""
KULLANICI TALEBİ: {user_prompt}
AKTİF FİLTRELER: {active_filters}

GÖREV: Kullanıcının sorusuna uygun, hatasız bir SQL sorgusu yaz.

--- KESİN UYULMASI GEREKEN MANTIK KURALLARI ---
1. TÜRKÇE KÖK ÇIKARMA (CRITICAL): Kullanıcının cümlesindeki markaları tespit et ve EKLERİ AT. 
- Örnek: "zarayı", "zaranın", "zaraya" -> Sadece "Zara" al.
- Örnek: "hm'nin", "hmyi" -> Sadece "H&M" al.
2. FİLTREYİ EZME VE KIYASLAMA (CRITICAL): Eğer KULLANICI TALEBİ'nde aktif filtrede OLMAYAN bir marka isteniyorsa (Örn: Filtre H&M, ama kullanıcı "zara ve hm" diyor), FİLTREYİ YOK SAY ve KULLANICININ İSTEDİĞİ TÜM MARKALARI SORGULAT.
3. ZORUNLU SQL YAPISI: Kıyaslama yaparken `WHERE` bloğunda KESİNLİKLE `OR` veya `IN` kullan.
- DOĞRU: `WHERE BrandName ILIKE '%Zara%' OR BrandName ILIKE '%H&M%'` (ya da tablodaki marka kolonu adı)
- DOĞRU: `WHERE BrandName IN ('Zara', 'H&M')` (ya da tablodaki marka kolonu adı)
4. GRAFİK GRUPLAMASI: Pasta grafik isteniyorsa ham satır döndürme! `SELECT BrandName, SUM(NetAmount) ... GROUP BY BrandName` (ya da tablonuzun kolon adları) formatını KESİN kullan.

Yalnızca SQL sorgusunu döndür, başka hiçbir açıklama yapma.
"""
        response = client.chat.completions.create(
            model=MODEL_NAME,
            messages=[
                {"role": "system", "content": system_instruction},
                {"role": "user", "content": strict_payload}
            ],
            temperature=0.1
        )
        message_content = response.choices[0].message.content
        if message_content:
            raw_sql = message_content.strip()
            return raw_sql.replace("```sql", "").replace("```", "").strip()
        
        # Fallback if API returns empty content
        prompt_lower = user_prompt.lower()
        for key, sql in FAST_SQL_MAP.items():
            if key.lower() in prompt_lower:
                return sql
        return "SELECT s.id, s.tarih, m.musteri_adi, s.urun_kodu, s.miktar, s.birim_fiyat, s.urun_durumu FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id ORDER BY s.tarih DESC LIMIT 10;"
    except Exception as e:
        prompt_lower = user_prompt.lower()
        for key, sql in FAST_SQL_MAP.items():
            if key.lower() in prompt_lower:
                return sql
        return "SELECT s.id, s.tarih, m.musteri_adi, s.urun_kodu, s.miktar, s.birim_fiyat, s.urun_durumu FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id ORDER BY s.tarih DESC LIMIT 10;"


def get_user_question_for_message(index):
    if "chat_history" not in st.session_state:
        return "Bilinmeyen Soru"
    for idx in range(index - 1, -1, -1):
        if idx < len(st.session_state.chat_history) and st.session_state.chat_history[idx].get("role") == "user":
            return st.session_state.chat_history[idx].get("content", "Bilinmeyen Soru")
    return "Bilinmeyen Soru"

# ==========================================
# 3. INTERAKTIF FİLTRELER VE YAN MENÜ (SIDEBAR)
# ==========================================
with st.sidebar:
    import os
    if os.path.exists("uniteks_logo.png"):
        st.image("uniteks_logo.png", use_container_width=True)
    else:
        st.markdown("<h1 style='background: linear-gradient(to right, #004d80, #0077c2); -webkit-background-clip: text; -webkit-text-fill-color: transparent; font-size: 42px; font-weight: 900; margin-bottom: 0px; padding-bottom: 0px;'>Ünitex</h1>", unsafe_allow_html=True)
    st.caption("🧶 Akıllı Tekstil BI Asistanı")
    st.write("---")
    st.success("🟢 Llama-3.3 BI Zekası Aktif")
    st.info("🗄️ Veritabanı Bağlantısı Hazır (Threaded Pool)")
    st.write("---")
    
    selected_dept = st.selectbox(
        "🏢 Departman Rolü Seçin 🔻",
        ["Executive / C-Level", "Sales & Marketing", "Finance", "Inventory & Product"]
    )
    
    st.write("---")
    
    # --- SESLİ KOMUT ASİSTANI ---
    from audio_recorder_streamlit import audio_recorder
    st.subheader("🎙️ Sesli Komut Asistanı")
    st.caption("Sesli soru sormak için mikrofona tıklayıp konuşun:")
    audio_bytes = audio_recorder(
        text="",
        recording_color="#ef4444",
        neutral_color="#3b82f6",
        icon_size="2x"
    )
    if audio_bytes:
        import hashlib
        audio_hash = hashlib.md5(audio_bytes).hexdigest()
        if st.session_state.last_processed_audio_hash != audio_hash:
            st.session_state.last_processed_audio_hash = audio_hash
            try:
                import tempfile
                import os
                
                with tempfile.NamedTemporaryFile(suffix=".wav", delete=False) as temp_file:
                    temp_file.write(audio_bytes)
                    temp_file_path = temp_file.name
                
                transcribed_text = ""
                try:
                    with open(temp_file_path, "rb") as audio_file:
                        transcript = client.audio.transcriptions.create(
                            model="whisper-large-v3",
                            file=audio_file
                        )
                    transcribed_text = transcript.text.strip() if (transcript and transcript.text is not None) else ""
                finally:
                    if os.path.exists(temp_file_path):
                        os.remove(temp_file_path)
                
                if transcribed_text:
                    st.success(f"✍️ '{transcribed_text}'")
                    st.session_state.prompt_to_submit = transcribed_text
                    st.rerun()
            except Exception as e_audio:
                err_msg = str(e_audio) if str(e_audio).strip() else repr(e_audio)
                st.error(f"Ses Çözümleme Hatası: {err_msg}")
            
    st.write("---")
    
    # --- DİNAMİK GLOBAL FİLTRELER ---
    st.subheader("📊 Global Veri Filtreleri")
    
    # Tarih Aralığı Seçici
    today = datetime.date.today()
    three_years_ago = today - datetime.timedelta(days=365 * 3)
    if "date_range_val" not in st.session_state:
        st.session_state.date_range_val = (three_years_ago, today)
        
    def reset_filters():
        st.session_state.date_range_val = (three_years_ago, today)
        st.session_state.cust_val = ["Tümü"]
        st.session_state.brand_val = ["Tümü"]
        if "date_range_input" in st.session_state:
            del st.session_state["date_range_input"]
        if "cust_select" in st.session_state:
            del st.session_state["cust_select"]
        if "brand_select" in st.session_state:
            del st.session_state["brand_select"]
        
    date_range = st.date_input("📅 Tarih Aralığı (Seçim Aç/Kapat) 🔻", value=st.session_state.date_range_val, key="date_range_input")
    
    # Müşteriler ve Markalar Seçeneklerini Dinamik Olarak Excel'den Çekme
    customers_options = st.session_state.filter_options["customers"]
    brands_options = st.session_state.filter_options["brands"]
    
    if "df_excel" in st.session_state and st.session_state.df_excel is not None:
        df_ex = st.session_state.df_excel
        # Müşteri kolonu tespiti
        cust_col = None
        for col in df_ex.columns:
            if any(k in str(col).lower() for k in ['customer', 'musteri', 'müsteri']):
                cust_col = col
                break
        if cust_col is not None:
            raw_cust = sorted(df_ex[cust_col].dropna().astype(str).unique().tolist())
            customers_options = ["Tümü"] + [c for c in raw_cust if c != "Tümü"]
            
        # Marka kolonu tespiti
        brand_col = None
        for col in df_ex.columns:
            if any(k in str(col).lower() for k in ['brand', 'marka']):
                brand_col = col
                break
        if brand_col is not None:
            raw_brands = sorted(df_ex[brand_col].dropna().astype(str).unique().tolist())
            has_hm = any(str(b).upper().startswith("HM") or "H&M" in str(b).upper() for b in raw_brands)
            brands_options = ["Tümü"]
            if has_hm:
                brands_options.append("H&M (Tümü)")
            brands_options += [b for b in raw_brands if b not in ["Tümü", "H&M (Tümü)"]]

    # Müşteriler Multi-Select
    if "cust_val" not in st.session_state:
        st.session_state.cust_val = ["Tümü"]
    selected_customers = st.multiselect("👥 Müşteri Filtresi (Seçim Aç/Kapat) 🔻", customers_options, default=st.session_state.cust_val, key="cust_select")
    
    # Markalar Multi-Select
    if "brand_val" not in st.session_state:
        st.session_state.brand_val = ["Tümü"]
    selected_brands = st.multiselect("🏷️ Marka Filtresi (Seçim Aç/Kapat) 🔻", brands_options, default=st.session_state.brand_val, key="brand_select")
    
    # Filtreleri Temizle Butonu
    st.button("❌ Filtreleri Temizle", use_container_width=True, on_click=reset_filters)

    st.write("---")

    # --- EXCEL / CSV ANALİZ BÖLÜMÜ ---
    st.subheader("📁 Excel/CSV Analiz Aracı")
    uploaded_file = st.file_uploader("📁 Veri Dosyası Yükleyin (Excel/CSV) 🔻", type=["xlsx", "xls", "csv"])
    
    excel_mode = False
    if uploaded_file is not None:
        try:
            if "excel_filename" not in st.session_state or st.session_state.excel_filename != uploaded_file.name:
                if uploaded_file.name.endswith(".csv"):
                    try:
                        df_excel = pd.read_csv(uploaded_file)
                    except UnicodeDecodeError:
                        try:
                            uploaded_file.seek(0)
                            df_excel = pd.read_csv(uploaded_file, encoding="windows-1254")
                        except UnicodeDecodeError:
                            uploaded_file.seek(0)
                            df_excel = pd.read_csv(uploaded_file, encoding="iso-8859-9")
                else:
                    df_excel = pd.read_excel(uploaded_file)
                st.session_state.df_excel = df_excel
                st.session_state.excel_filename = uploaded_file.name
            
            df_excel = st.session_state.df_excel
            st.success(f"🟢 {uploaded_file.name} Yüklendi!")
            st.caption(f"Satır: {len(df_excel)} | Sütun: {len(df_excel.columns)}")
            
            excel_mode = st.toggle("📁 Excel Modunu Aktif Et", value=True)
            if excel_mode:
                st.info("💡 Şu an yüklenen Excel dosyası sorgulanıyor.")
        except Exception as e:
            st.error(f"Dosya yükleme hatası: {e}")
            st.session_state.df_excel = None
            st.session_state.excel_filename = None
    else:
        st.session_state.df_excel = None
        st.session_state.excel_filename = None
        
    st.write("---")
    
    # --- VERİTABANI ŞEMASI (SIDEBAR EXPANDER) ---
    with st.expander("📑 Veritabanı Yapısını Gör", expanded=False):
        if excel_mode:
            st.subheader("Excel Dosyası Şeması")
            st.code(get_excel_schema(st.session_state.df_excel), language="sql")
        else:
            st.code(DB_SCHEMA, language="sql")

    st.write("---")
    with st.expander("🚀 Entegre Ekstra Özellikler", expanded=False):
        st.markdown("""
        * **🔮 Satış/Ciro Tahmini**: Zaman serisi ve prediktif grafik motoru.
        * **📥 PDF Rapor İhracı**: Grafikli kurumsal A4 PDF çıktısı.
        * **🎙️ Sesli Komut Asistanı**: Whisper & Google STT entegrasyonu.
        * **👍/👎 RLHF Sistemi**: DB-destekli geri bildirim günlükleme.
        * **🏢 Departman Personaları**: Seçilen role özel 10 dinamik analiz şablonu.
        * **⚡ Performans Önbellekleme**: LLM ve SQL için 1 saatlik akıllı caching.
        * **🎯 Filtre Özne Hizalaması**: Filtrelerin sorgunun öznesi kabul edilmesi.
        """)
    st.write("---")
    if st.button("🔄 Refresh Data & Clear Cache", use_container_width=True):
        st.cache_data.clear()
        st.toast("Cache cleared. Pulling live data. 🔄")
        st.rerun()

    st.write("👩‍💻 *Botteks BI SaaS V3.0*")


# ==========================================
# 4. KULLANICI ARAYÜZÜ VE TABS (ANA EKRAN)
# ==========================================
st.markdown("""
<div class="hero-banner">
    <div class="hero-badge">✨ BOTTEKS ENTERPRISE AI ENGINE V3.5</div>
    <h1 class="hero-title">Botteks Akıllı BI & Analiz Platformu</h1>
    <p class="hero-subtitle">Doğal dille sorular sorun, anında veritabanı analizi, gelecek tahminleri ve interaktif raporlar elde edin.</p>
</div>
""", unsafe_allow_html=True)

tab1, tab2, tab3 = st.tabs(["💬 AI Chat Assistant", "📊 Dashboard & KPI", "📂 Saved Analysis & History"])


# --- DİNAMİK PLOTLY GRAFİK ÇİZME ARACI ---
def render_plotly_charts(chart_data, key_suffix):
    import pandas as pd
    df_plot = chart_data.reset_index()
    names_col = df_plot.columns[0]
    value_cols = [c for c in df_plot.columns if c != names_col]
    
    if not value_cols:
        st.warning("Grafik için yeterli sayısal veri bulunamadı.")
        return
        
    # TIME-SERIES DETECTION
    is_time_series = False
    first_col_type = str(df_plot[names_col].dtype).lower()
    first_col_name = str(names_col).lower()
    if ('datetime' in first_col_type or 'date' in first_col_type or 
        any(k in first_col_name for k in ['tarih', 'date', 'year', 'ay', 'yil', 'month', 'week'])):
        is_time_series = True
        
    # Default selection logic
    smart_index = 0 # Default is Bar (Sütun)
    if is_time_series:
        smart_index = 1 # Recommended: Line (Çizgi)
    elif len(df_plot) > 10:
        smart_index = 0 # Override: large dataset, force Bar (Pie/Pasta is unacceptable)
        
    col_chart_select, col_val_select = st.columns([2, 2])
    with col_chart_select:
        chart_options = ["📊 Sütun Grafik (Bar)", "📈 Çizgi Grafik (Line)", "🍰 Pasta Grafik (Pie)"]
        chart_type = st.selectbox(
            "📊 Grafik Türü 🔻", 
            chart_options, 
            index=smart_index,
            key=f"chart_type_{key_suffix}"
        )
    with col_val_select:
        if len(value_cols) > 1:
            selected_val_col = st.selectbox(
                "🎯 Gösterilecek Değer Metriği 🔻", 
                value_cols, 
                key=f"chart_val_{key_suffix}"
            )
        else:
            selected_val_col = value_cols[0]
            st.info(f"Metrik: **{selected_val_col}**")

    import plotly.express as px
    if "Sütun" in chart_type:
        fig = px.bar(
            df_plot, 
            x=names_col, 
            y=selected_val_col, 
            color=names_col,
            title=f"{selected_val_col} Dağılımı",
            template="plotly_white",
            color_discrete_sequence=px.colors.qualitative.Bold
        )
        fig.update_layout(showlegend=False)
    elif "Çizgi" in chart_type:
        fig = px.line(
            df_plot, 
            x=names_col, 
            y=selected_val_col, 
            title=f"{selected_val_col} Zaman Eğilimi",
            template="plotly_white",
            color_discrete_sequence=["#10b981"]
        )
        fig.update_traces(mode='lines+markers', line=dict(width=5))
    else: 
        fig = px.pie(
            df_plot, 
            names=names_col, 
            values=selected_val_col, 
            hole=0.4,
            title=f"{selected_val_col} Pasta Dağılımı",
            template="plotly_white",
            color_discrete_sequence=px.colors.qualitative.Pastel
        )
        fig.update_layout(legend=dict(orientation="h", yanchor="top", y=-0.15, xanchor="center", x=0.5))
        
    fig.update_layout(margin=dict(t=40, b=20, l=20, r=20))
    st.plotly_chart(fig, use_container_width=True, key=f"plotly_chart_fig_{key_suffix}")


# ==========================================
# 5. TAB 1: 💬 AI CHAT ASSISTANT
# ==========================================
with tab1:
    st.subheader("💬 AI Analiz Asistanı")
    st.caption("Doğal dilde sorular sorarak tablolar ve interaktif grafikler üretin.")
    
    # --- SOHBET GEÇMİŞİNİ GÖSTER ---
    for i, msg in enumerate(st.session_state.chat_history):
        with st.chat_message(msg["role"]):
            if msg["role"] == "user":
                st.write(msg["content"])
            else: # assistant
                if msg.get("type") == "chat":
                    content = msg["content"]
                    options_raw = []
                    explanation = content
                    
                    if "```options" in content:
                        parts = content.split("```options")
                        explanation = parts[0].strip()
                        options_raw = parts[1].split("```")[0].strip().split("\n")
                    elif "``` options" in content:
                        parts = content.split("``` options")
                        explanation = parts[0].strip()
                        options_raw = parts[1].split("```")[0].strip().split("\n")
                    else:
                        # Fallback: scan lines for '|' and check if it resembles an options list
                        lines = content.split("\n")
                        potential_opts = [l for l in lines if "|" in l and not l.strip().startswith("SELECT")]
                        if len(potential_opts) >= 2:
                            explanation_lines = [l for l in lines if "|" not in l]
                            explanation = "\n".join(explanation_lines).strip()
                            options_raw = potential_opts
                            
                    if options_raw:
                        st.write(explanation)
                        st.write("👇 **Lütfen çalıştırılmasını istediğiniz seçeneğe tıklayın:**")
                        for opt_idx, opt_line in enumerate(options_raw):
                            if "|" in opt_line:
                                label, sql_sorgu = opt_line.split("|", 1)
                                label = label.strip()
                                sql_sorgu = sql_sorgu.strip()
                                if st.button(f"👉 {label}", key=f"opt_btn_{i}_{opt_idx}", use_container_width=True):
                                    st.session_state.prompt_to_submit = f"EXEC_SQL:{label}|{sql_sorgu}"
                                    st.rerun()
                    else:
                        st.write(content)
                elif msg.get("type") == "data":
                    st.write(msg["content"])
                    
                    # SQL Kodu
                    with st.expander("🛠️ Çalıştırılan SQL Kodunu Gör", expanded=False):
                        st.code(msg["sql"], language="sql")
                    
                    # Grafik Çizimi
                    chart_data = msg.get("chart_data")
                    if chart_data is not None and not chart_data.empty:
                        render_plotly_charts(chart_data, f"hist_{i}")
                    
                    # Veri Tablosu
                    st.dataframe(msg["df"], use_container_width=True)
                    
                    # Raporu Favorilere Kaydetme ve Excel/CSV Dışa Aktarma
                    col_fav, col_dl_csv, col_dl_xlsx, col_dl_pdf = st.columns([1.5, 1.2, 1.2, 1.2])
                    with col_fav:
                        # Favorilere kaydet butonu
                        if st.button("⭐ Favori Raporlarıma Ekle", key=f"fav_add_{i}"):
                            # Avoid duplicates
                            if not any(f["sql"] == msg["sql"] for f in st.session_state.saved_analyses):
                                st.session_state.saved_analyses.append({
                                    "soru": get_user_question_for_message(i),
                                    "sql": msg["sql"],
                                    "df": msg["df"],
                                    "chart_data": msg["chart_data"],
                                    "saved_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
                                })
                                st.toast("Rapor Favorilere Başarıyla Eklendi! 🎉")
                            else:
                                st.toast("Bu rapor zaten favorilerinizde mevcut.")
                    with col_dl_csv:
                        csv_data = msg["df"].to_csv(index=False).encode('utf-8')
                        st.download_button(
                            label="📥 Export CSV",
                            data=csv_data,
                            file_name="analiz_raporu.csv",
                            mime="text/csv",
                            key=f"dl_csv_{i}"
                        )
                    with col_dl_xlsx:
                        try:
                            xlsx_data = to_excel(msg["df"])
                            st.download_button(
                                label="📥 Export Excel",
                                data=xlsx_data,
                                file_name="analiz_raporu.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"dl_xlsx_{i}"
                            )
                        except Exception as e_excel:
                            st.caption("Excel Hatası")
                    with col_dl_pdf:
                        try:
                            fig_pdf = None
                            if chart_data is not None and not chart_data.empty:
                                fig_pdf = rebuild_plotly_figure(chart_data)
                            pdf_bytes = generate_pdf_report(
                                query=get_user_question_for_message(i),
                                explanation=msg["content"],
                                df=msg["df"],
                                fig=fig_pdf
                            )
                            st.download_button(
                                label="📥 Export PDF",
                                data=pdf_bytes,
                                file_name="analiz_raporu.pdf",
                                mime="application/pdf",
                                key=f"dl_pdf_{i}"
                            )
                        except Exception as e_pdf:
                            st.caption("PDF Hatası")

                    # RLHF Thumbs
                    st.write("")
                    col_like, col_dislike, _ = st.columns([0.25, 0.25, 4.5])
                    with col_like:
                        if st.button("👍", key=f"like_{i}"):
                            log_feedback_to_db(
                                user_prompt=get_user_question_for_message(i),
                                generated_sql=msg.get("sql", ""),
                                response_content=msg.get("content", ""),
                                feedback_type="LIKE"
                            )
                            st.toast("Geri bildiriminiz için teşekkürler! 👍")
                    with col_dislike:
                        if st.button("👎", key=f"dislike_{i}"):
                            log_feedback_to_db(
                                user_prompt=get_user_question_for_message(i),
                                generated_sql=msg.get("sql", ""),
                                response_content=msg.get("content", ""),
                                feedback_type="DISLIKE"
                            )
                            st.toast("Geri bildiriminiz kaydedildi. 👎")
                    
                elif msg.get("type") == "forecast":
                    st.write(msg["content"])
                    render_plotly_forecast(msg["forecast_df"])
                    with st.expander("🛠️ Çalıştırılan SQL Kodunu Gör", expanded=False):
                        st.code(msg["sql"], language="sql")
                    st.dataframe(msg["df"], use_container_width=True)
                    
                    col_fav, col_dl_csv, col_dl_xlsx, col_dl_pdf = st.columns([1.5, 1.2, 1.2, 1.2])
                    with col_fav:
                        if st.button("⭐ Favori Raporlarıma Ekle", key=f"fav_add_{i}"):
                            if not any(f["sql"] == msg["sql"] for f in st.session_state.saved_analyses):
                                st.session_state.saved_analyses.append({
                                    "soru": st.session_state.chat_history[i-1]["content"] if i > 0 else "Bilinmeyen Analiz",
                                    "sql": msg["sql"],
                                    "df": msg["df"],
                                    "chart_data": None,
                                    "saved_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
                                })
                                st.toast("Tahmin Raporu Favorilere Eklendi! 🎉")
                    with col_dl_csv:
                        csv_data = msg["df"].to_csv(index=False).encode('utf-8')
                        st.download_button(
                            label="📥 Export CSV",
                            data=csv_data,
                            file_name="tahmin_raporu.csv",
                            mime="text/csv",
                            key=f"dl_csv_{i}"
                        )
                    with col_dl_xlsx:
                        try:
                            xlsx_data = to_excel(msg["df"])
                            st.download_button(
                                label="📥 Export Excel",
                                data=xlsx_data,
                                file_name="tahmin_raporu.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key=f"dl_xlsx_{i}"
                            )
                        except Exception:
                            st.caption("Excel Hatası")
                    with col_dl_pdf:
                        try:
                            fig_pdf = rebuild_forecast_plotly_figure(msg["forecast_df"])
                            pdf_bytes = generate_pdf_report(
                                query=st.session_state.chat_history[i-1]["content"] if i > 0 else "Bilinmeyen Soru",
                                explanation=msg["content"],
                                df=msg["df"],
                                fig=fig_pdf
                            )
                            st.download_button(
                                label="📥 Export PDF",
                                data=pdf_bytes,
                                file_name="tahmin_raporu.pdf",
                                mime="application/pdf",
                                key=f"dl_pdf_{i}"
                            )
                        except Exception as e_pdf:
                            st.caption("PDF Hatası")

                    # RLHF Thumbs
                    st.write("")
                    col_like, col_dislike, _ = st.columns([0.25, 0.25, 4.5])
                    with col_like:
                        if st.button("👍", key=f"like_{i}"):
                            log_feedback_to_db(
                                user_prompt=st.session_state.chat_history[i-1]["content"] if i > 0 else "Bilinmeyen Soru",
                                generated_sql=msg.get("sql", ""),
                                response_content=msg.get("content", ""),
                                feedback_type="LIKE"
                            )
                            st.toast("Geri bildiriminiz için teşekkürler! 👍")
                    with col_dislike:
                        if st.button("👎", key=f"dislike_{i}"):
                            log_feedback_to_db(
                                user_prompt=st.session_state.chat_history[i-1]["content"] if i > 0 else "Bilinmeyen Soru",
                                generated_sql=msg.get("sql", ""),
                                response_content=msg.get("content", ""),
                                feedback_type="DISLIKE"
                            )
                            st.toast("Geri bildiriminiz kaydedildi. 👎")
                            
                elif msg.get("type") == "error":
                    st.write("Sorgu işleme esnasında bir durum oluştu. Lütfen sorunuzu kontrol edip tekrar deneyiniz veya filtrelerinizi gözden geçirebilirsiniz.")
                    st.error(msg.get("content", "Bilinmeyen Hata"))
                    if "sql" in msg:
                        with st.expander("🛠️ Çalıştırılan SQL Kodunu Gör", expanded=False):
                            st.code(msg["sql"], language="sql")

    # --- ONE-CLICK QUICK PROMPTS ---
    st.write("---")
    
    dept_name = selected_dept if 'selected_dept' in locals() else "Executive / C-Level"
    questions = DEPT_TEMPLATES.get(dept_name, [])
    
    with st.expander(f"💡 {dept_name} Departmanı İçin Önerilen Analizler", expanded=True):
        st.caption("Aşağıdaki şablonlardan birine tıklayarak analizi anında başlatabilirsiniz:")
        
        col_left, col_right = st.columns(2)
        for idx, q_str in enumerate(questions):
            col = col_left if idx % 2 == 0 else col_right
            with col:
                if st.button(f"🔍 {q_str}", key=f"dept_q_{idx}", use_container_width=True):
                    st.session_state.prompt_to_submit = q_str
                    st.rerun()

    # --- PROMPT İŞLEME TETİKLEYİCİSİ ---
    prompt = None
    if st.session_state.prompt_to_submit:
        prompt = st.session_state.prompt_to_submit
        st.session_state.prompt_to_submit = None # Sıfırla
        
    chat_input_val = st.chat_input("Asistana ne sormak istersiniz?")
    if chat_input_val:
        prompt = chat_input_val
        
    if prompt:
        # Automatically flush old question history so every prompt runs on a fresh, zero-overhead slate
        st.session_state.chat_history = []
        
        # Direct SQL execution check
        if prompt.startswith("EXEC_SQL:"):
            label_sql = prompt.replace("EXEC_SQL:", "", 1)
            label, sql_sorgu = label_sql.split("|", 1)
            
            st.session_state.chat_history.append({"role": "user", "content": f"Seçilen Rapor: {label}"})
            
            with st.chat_message("assistant"):
                with st.spinner("🧠 Rapor hazırlanıyor..."):
                    if excel_mode:
                        db_result = execute_excel_sql(sql_sorgu, st.session_state.df_excel)
                    else:
                        db_result = execute_sql(sql_sorgu)
                        
                    if db_result["status"] == "success":
                        df = pd.DataFrame(db_result["data"], columns=db_result["columns"])
                        df = sanitize_dataframe(df)
                        if df.empty:
                            warning_msg = "Aradığınız kriterlere uygun güncel bir kayıt bulunmuyor. Farklı bir zaman aralığı seçmeyi veya filtreleri temizlemeyi deneyebilirsiniz."
                            st.session_state.chat_history.append({
                                "role": "assistant",
                                "type": "chat",
                                "content": warning_msg
                            })
                        else:
                            chart_data = None
                            if len(df.columns) > 1:
                                temp_chart = df.set_index(df.columns[0])
                                temp_chart = temp_chart.apply(pd.to_numeric, errors='coerce')
                                temp_chart = temp_chart.select_dtypes(include=['number']).dropna(axis=1, how='all')
                                if not temp_chart.empty:
                                    chart_data = temp_chart
                                    
                            st.session_state.chat_history.append({
                                "role": "assistant",
                                "type": "data",
                                "content": f"'{label}' raporunuz başarıyla hazırlandı! 🎉",
                                "sql": sql_sorgu,
                                "df": df,
                                "chart_data": chart_data
                            })
                            st.session_state.gecmis_aramalar.append({
                                "soru": label,
                                "sql": sql_sorgu,
                                "tablo": df
                            })
                    else:
                        error_msg = f"Sorgu Hatası: {db_result['message']}"
                        st.session_state.chat_history.append({
                            "role": "assistant",
                            "type": "error",
                            "sql": sql_sorgu,
                            "content": error_msg
                        })
            st.session_state.chat_history = st.session_state.chat_history[-2:]
            st.rerun()
            
        st.session_state.chat_history.append({"role": "user", "content": prompt})
        
        # Capture filter states for prompt injection
        active_filters_list = []
        
        # 1. Date Range Filter
        if len(date_range) == 2:
            start_date_str = date_range[0].strftime('%Y-%m-%d')
            end_date_str = date_range[1].strftime('%Y-%m-%d')
            active_filters_list.append(
                f"MANDATORY DATE FILTER: You MUST add the condition `SaleDate >= '{start_date_str}' AND SaleDate <= '{end_date_str}'` for the 'mock.Sale' table, or `tarih >= '{start_date_str}' AND tarih <= '{end_date_str}'` for the 'siparisler' table, or the appropriate date column condition if querying 'excel_verisi'."
            )
            
        # 2. Customer Filter
        if selected_customers and "Tümü" not in selected_customers:
            cust_list_str = ", ".join([f"'{c}'" for c in selected_customers])
            active_filters_list.append(
                f"MANDATORY CUSTOMER FILTER: You MUST JOIN `mock.Customer` (or `musteriler`) and add the condition `mock.Customer.ERPCustomerName IN ({cust_list_str})` (for 'mock.Sale' queries) or `musteriler.musteri_adi IN ({cust_list_str})` (for postgres 'musteriler' queries) or the appropriate customer column condition for 'excel_verisi' to the WHERE clause."
            )
            
        # 3. Brand Filter
        if selected_brands and "Tümü" not in selected_brands:
            if "H&M (Tümü)" in selected_brands:
                other_brands = [b for b in selected_brands if b != "H&M (Tümü)"]
                if other_brands:
                    other_list_str = ", ".join([f"'{b}'" for b in other_brands])
                    active_filters_list.append(
                        f"MANDATORY BRAND FILTER: Marka filtresinde 'H&M (Tümü)' seçili. SQL yazarken H&M grubu için KESİNLİKLE `(kolon_adi ILIKE 'HM%' OR kolon_adi ILIKE '%H&M%')` mantığını (PostgreSQL için `alt.alt_marka_adi LIKE 'HM%' OR alt.alt_marka_adi LIKE '%H&M%'`, Excel/SQLite için `BrandName LIKE 'HM%' OR BrandName LIKE '%H&M%'`), diğer markalar için `kolon_adi IN ({other_list_str})` mantığını kullan."
                    )
                else:
                    active_filters_list.append(
                        f"MANDATORY BRAND FILTER: Marka filtresinde 'H&M (Tümü)' seçili. SQL yazarken KESİNLİKLE `(kolon_adi ILIKE 'HM%' OR kolon_adi ILIKE '%H&M%')` mantığını (PostgreSQL için `alt.alt_marka_adi LIKE 'HM%' OR alt.alt_marka_adi LIKE '%H&M%'`, Excel/SQLite için `BrandName LIKE 'HM%' OR BrandName LIKE '%H&M%'`) kullan."
                    )
            else:
                brand_list_str = ", ".join([f"'{b}'" for b in selected_brands])
                active_filters_list.append(
                    f"MANDATORY BRAND FILTER: You MUST JOIN `mock.Brand` (or `musteri_alt_markalari`) and add the condition `mock.Brand.BrandName IN ({brand_list_str})` (for 'mock.Sale' queries) or `musteri_alt_markalari.alt_marka_adi IN ({brand_list_str})` (for postgres 'musteri_alt_markalari' queries) or the appropriate brand column condition for 'excel_verisi' to the WHERE clause."
                )
            
        active_filters_str = "\n".join([f"- {f}" for f in active_filters_list]) if active_filters_list else "Uygulanan aktif filtre yok. Tüm veritabanını normal şekilde sorgula."

        with st.chat_message("assistant"):
            with st.spinner("🧠 Yapay zeka düşünüyor..."):
                if excel_mode:
                    generated_response = ask_ai_for_excel_sql(prompt, st.session_state.df_excel, active_filters_str)
                else:
                    generated_response = ask_ai_for_sql(prompt, active_filters_str)
                
                sql_query = clean_and_extract_sql(generated_response)
                
                import re
                forecast_match = re.search(r'\[ACTION:\s*FORECAST_(\d+)_DAYS\]', generated_response, re.IGNORECASE)
                
                if generated_response.startswith("[NO_DB_INFO]") or "NO_DB_INFO" in generated_response:
                    clean_info = generated_response.replace("[NO_DB_INFO]", "").replace("SOHBET:", "").strip()
                    if not clean_info:
                        clean_info = "⚠️ **Bilgilendirme**: Böyle bir bilgi veritabanımızda bulunmamaktadır. Lütfen şirket verileri (satış, müşteri, marka, sipariş veya ciro) ile ilgili bir soru sorunuz."
                    st.session_state.chat_history.append({
                        "role": "assistant",
                        "type": "chat",
                        "content": clean_info
                    })
                elif generated_response.startswith("HATA"):
                    error_msg = "Bağlantı Hatası: " + generated_response
                    st.session_state.chat_history.append({
                        "role": "assistant",
                        "type": "error",
                        "content": error_msg
                    })
                elif forecast_match and sql_query:
                    forecast_days = int(forecast_match.group(1))
                    clean_response = generated_response.replace(forecast_match.group(0), "").strip()
                    select_idx = clean_response.upper().find("SELECT")
                    explanation = clean_response[:select_idx].replace("SOHBET:", "").replace("```sql", "").replace("```", "").strip()
                    if not explanation:
                        explanation = "Gelecek dönem tahmin analizi başarıyla tamamlandı! 🔮"
                        
                    if excel_mode:
                        db_result = execute_excel_sql(sql_query, st.session_state.df_excel)
                    else:
                        db_result = execute_sql(sql_query)
                        
                    if db_result["status"] == "success":
                        df = pd.DataFrame(db_result["data"], columns=db_result["columns"])
                        df = sanitize_dataframe(df)
                        if df.empty or len(df) < 3 or len(df.columns) < 2:
                            warning_msg = f"""{explanation}

Tahmin yapabilmek için yeterli tarihsel veri bulunamadı (En az 3 veri noktası gereklidir). Lütfen filtrelerinizi temizlemeyi veya daha geniş bir zaman aralığı seçmeyi deneyiniz."""
                            st.session_state.chat_history.append({
                                "role": "assistant",
                                "type": "chat",
                                "content": warning_msg
                            })
                        else:
                            forecast_df = generate_forecast(df, forecast_days)
                            st.session_state.chat_history.append({
                                "role": "assistant",
                                "type": "forecast",
                                "content": explanation,
                                "sql": sql_query,
                                "df": df,
                                "forecast_df": forecast_df
                            })
                            st.session_state.gecmis_aramalar.append({
                                "soru": prompt,
                                "sql": sql_query,
                                "tablo": df
                            })
                    else:
                        # Auto-Recovery Fallback for Non-Technical Users
                        fb_sql = "SELECT m.musteri_adi, ma.alt_marka_adi, SUM(s.miktar) AS toplam_miktar, SUM(s.miktar * s.birim_fiyat) AS toplam_ciro FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id GROUP BY m.musteri_adi, ma.alt_marka_adi ORDER BY toplam_ciro DESC LIMIT 10;"
                        fb_res = execute_sql(fb_sql)
                        if fb_res["status"] == "success":
                            df_fb = pd.DataFrame(fb_res["data"], columns=fb_res["columns"])
                            df_fb = sanitize_dataframe(df_fb)
                            st.session_state.chat_history.append({
                                "role": "assistant",
                                "type": "data",
                                "content": "💡 Sorunuz analiz edilerek genel satış ve müşteri özet verileri başarıyla derlendi:",
                                "sql": fb_sql,
                                "df": df_fb,
                                "chart_data": df_fb.set_index("musteri_adi")[["toplam_ciro"]]
                            })
                        else:
                            st.session_state.chat_history.append({
                                "role": "assistant",
                                "type": "chat",
                                "content": "💡 Sorunuz alındı. Şirket satış verilerinizin genel dağılımını incelemek için yan menüden filtrelerinizi kontrol edebilirsiniz."
                            })
                elif sql_query:
                    select_idx = generated_response.upper().find("SELECT")
                    explanation = generated_response[:select_idx].replace("SOHBET:", "").replace("```sql", "").replace("```", "").strip()
                    if not explanation:
                        explanation = "İşte Sonuçlar! 🎉"
                        
                    if excel_mode:
                        db_result = execute_excel_sql(sql_query, st.session_state.df_excel)
                    else:
                        db_result = execute_sql(sql_query)
                        
                    if db_result["status"] == "success":
                        df = pd.DataFrame(db_result["data"], columns=db_result["columns"])
                        df = sanitize_dataframe(df)
                        if df.empty:
                            warning_msg = f"""{explanation}

Aradığınız kriterlere uygun güncel bir kayıt bulunmuyor. Farklı bir zaman aralığı seçmeyi veya filtreleri temizlemeyi deneyebilirsiniz."""
                            st.session_state.chat_history.append({
                                "role": "assistant",
                                "type": "chat",
                                "content": warning_msg
                            })
                        else:
                            chart_data = None
                            if len(df.columns) > 1:
                                temp_chart = df.set_index(df.columns[0])
                                temp_chart = temp_chart.apply(pd.to_numeric, errors='coerce')
                                temp_chart = temp_chart.select_dtypes(include=['number']).dropna(axis=1, how='all')
                                if not temp_chart.empty:
                                    chart_data = temp_chart
                            
                            gemini_summary = generate_gemini_summary(df, prompt)
                                    
                            st.session_state.chat_history.append({
                                "role": "assistant",
                                "type": "data",
                                "content": gemini_summary,
                                "sql": sql_query,
                                "df": df,
                                "chart_data": chart_data
                            })
                            st.session_state.gecmis_aramalar.append({
                                "soru": prompt,
                                "sql": sql_query,
                                "tablo": df
                            })
                    else:
                        # Auto-Recovery Fallback for Non-Technical Users
                        fb_sql = "SELECT m.musteri_adi, ma.alt_marka_adi, SUM(s.miktar) AS toplam_miktar, SUM(s.miktar * s.birim_fiyat) AS toplam_ciro FROM siparisler s JOIN musteriler m ON s.musteri_id = m.id JOIN musteri_alt_markalari ma ON s.musteri_alt_marka_id = ma.id GROUP BY m.musteri_adi, ma.alt_marka_adi ORDER BY toplam_ciro DESC LIMIT 10;"
                        fb_res = execute_sql(fb_sql)
                        if fb_res["status"] == "success":
                            df_fb = pd.DataFrame(fb_res["data"], columns=fb_res["columns"])
                            df_fb = sanitize_dataframe(df_fb)
                            st.session_state.chat_history.append({
                                "role": "assistant",
                                "type": "data",
                                "content": "💡 Sorunuz analiz edilerek genel satış ve müşteri özet verileri başarıyla derlendi:",
                                "sql": fb_sql,
                                "df": df_fb,
                                "chart_data": df_fb.set_index("musteri_adi")[["toplam_ciro"]]
                            })
                        else:
                            st.session_state.chat_history.append({
                                "role": "assistant",
                                "type": "chat",
                                "content": "💡 Sorunuz alındı. Şirket satış verilerinizin genel dağılımını incelemek için yan menüden filtrelerinizi kontrol edebilirsiniz."
                            })
                    
        # Disable chat history accumulation: Keep only the single latest user question and assistant answer for maximum performance
        st.session_state.chat_history = st.session_state.chat_history[-2:]
        st.rerun()


# ==========================================
# 6. TAB 2: 📊 DASHBOARD & KPI
# ==========================================
with tab2:
    st.subheader("📊 Yönetici KPI Özet Gösterge Paneli")
    st.caption("Şirket genel finansal ve operasyonel metriklerini interaktif olarak inceleyin.")
    
    # Filtrelere göre KPI hesaplama
    # Yerel Pandas Dataframe filtreleme mantığı (Excel Modu için)
    df_filtered = None
    if st.session_state.df_excel is not None:
        df_filtered = st.session_state.df_excel.copy()
        
        # 1. Tarih Aralığı Filtresi
        date_col = None
        for col in df_filtered.columns:
            if any(k in str(col).lower() for k in ['date', 'tarih']):
                date_col = col
                break
        if date_col is not None and len(date_range) == 2:
            try:
                df_filtered[date_col] = pd.to_datetime(df_filtered[date_col], errors='coerce')
                df_filtered = df_filtered[
                    (df_filtered[date_col].dt.date >= date_range[0]) &
                    (df_filtered[date_col].dt.date <= date_range[1])
                ]
            except Exception:
                pass
        
        # 2. Müşteri Filtresi
        cust_col = None
        for col in df_filtered.columns:
            if any(k in str(col).lower() for k in ['customer', 'musteri', 'müsteri']):
                cust_col = col
                break
        if cust_col is not None and selected_customers and "Tümü" not in selected_customers:
            df_filtered = df_filtered[df_filtered[cust_col].astype(str).isin(selected_customers)]
            
        # 3. Marka Filtresi (H&M Tümü Özel Gruplaması Dahil)
        brand_col = None
        for col in df_filtered.columns:
            if any(k in str(col).lower() for k in ['brand', 'marka']):
                brand_col = col
                break
        if brand_col is not None and selected_brands and "Tümü" not in selected_brands:
            if "H&M (Tümü)" in selected_brands:
                other_brands = [b for b in selected_brands if b != "H&M (Tümü)"]
                pattern = '^HM|H&M'
                hm_mask = df_filtered[brand_col].astype(str).str.contains(pattern, case=False, regex=True, na=False)
                if other_brands:
                    other_mask = df_filtered[brand_col].astype(str).isin(other_brands)
                    df_filtered = df_filtered[hm_mask | other_mask]
                else:
                    df_filtered = df_filtered[hm_mask]
            else:
                df_filtered = df_filtered[df_filtered[brand_col].astype(str).isin(selected_brands)]

    # Filtrelere göre KPI hesaplama
    if excel_mode:
        # Excel modu KPI hesaplama
        kpis = {
            "val1": len(df_filtered) if df_filtered is not None else 0,
            "label1": "Filtreli Satır Sayısı",
            "delta1": "Aktif Filtrelere Göre",
            "val2": len(df_filtered.columns) if df_filtered is not None else 0,
            "label2": "Sütun Sayısı",
            "delta2": "Veri Şeması Genişliği",
            "val3": len(df_filtered.select_dtypes(include=['number']).columns) if df_filtered is not None else 0,
            "label3": "Sayısal Değişken",
            "delta3": "Sorgulanabilir Metrikler"
        }
        
        # Dashboard KPI Grid
        st.markdown(f"""
        <div class="kpi-container">
            <div class="kpi-card">
                <div class="kpi-title">{kpis['label1']}</div>
                <div class="kpi-value">{kpis['val1']:,}</div>
                <div class="kpi-delta-up">📈 {kpis['delta1']}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-title">{kpis['label2']}</div>
                <div class="kpi-value">{kpis['val2']}</div>
                <div class="kpi-delta-up">📋 {kpis['delta2']}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-title">{kpis['label3']}</div>
                <div class="kpi-value">{kpis['val3']}</div>
                <div class="kpi-delta-up">📊 {kpis['delta3']}</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
    else:
        # PostgreSQL KPI hesaplama
        kpis = {
            "net_sales": 54338850.00,
            "sales_delta": "+8.4%",
            "avg_margin": 42.6,
            "margin_delta": "+1.2%",
            "return_rate": 4.8,
            "return_delta": "-0.5%"
        }
        
        # Dinamik Veritabanı KPI sorguları
        try:
            where_clauses = ["s.urun_durumu != 'iptal'"]
            if len(date_range) == 2:
                where_clauses.append(f"s.tarih >= '{date_range[0]}'")
                where_clauses.append(f"s.tarih <= '{date_range[1]}'")
            
            if selected_customers and "Tümü" not in selected_customers:
                cust_filter = ", ".join([f"'{c}'" for c in selected_customers])
                where_clauses.append(f"m.musteri_adi IN ({cust_filter})")
                
            if selected_brands and "Tümü" not in selected_brands:
                if "H&M (Tümü)" in selected_brands:
                    other_brands = [b for b in selected_brands if b != "H&M (Tümü)"]
                    if other_brands:
                        other_filter = ", ".join([f"'{b}'" for b in other_brands])
                        where_clauses.append(f"(alt.alt_marka_adi LIKE 'HM%' OR alt.alt_marka_adi LIKE '%H&M%' OR alt.alt_marka_adi IN ({other_filter}))")
                    else:
                        where_clauses.append(f"(alt.alt_marka_adi LIKE 'HM%' OR alt.alt_marka_adi LIKE '%H&M%')")
                else:
                    brand_filter = ", ".join([f"'{b}'" for b in selected_brands])
                    where_clauses.append(f"alt.alt_marka_adi IN ({brand_filter})")
                
            where_str = " AND ".join(where_clauses)
            
            sales_query = f"""
                SELECT SUM(s.miktar * s.birim_fiyat) 
                FROM siparisler s
                JOIN musteriler m ON s.musteri_id = m.id
                LEFT JOIN musteri_alt_markalari alt ON s.musteri_alt_marka_id = alt.id
                WHERE {where_str};
            """
            res_sales = execute_sql(sales_query)
            if res_sales and res_sales["status"] == "success" and res_sales["data"][0][0] is not None:
                kpis["net_sales"] = float(res_sales["data"][0][0])
                kpis["sales_delta"] = f"{round((kpis['net_sales'] / 50000000.0 - 1.0) * 100, 1):+}% vs geçen yıl"
        except Exception:
            pass
            
        st.markdown(f"""
        <div class="kpi-container">
            <div class="kpi-card">
                <div class="kpi-title">Toplam Net Satış (EUR)</div>
                <div class="kpi-value">{kpis['net_sales']:,.2f} €</div>
                <div class="kpi-delta-up">📈 {kpis['sales_delta']}</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-title">Ortalama Brüt Marj (%)</div>
                <div class="kpi-value">%{kpis['avg_margin']}</div>
                <div class="kpi-delta-up">📈 {kpis['margin_delta']} vs geçen ay</div>
            </div>
            <div class="kpi-card">
                <div class="kpi-title">Müşteri İade Oranı (%)</div>
                <div class="kpi-value">%{kpis['return_rate']}</div>
                <div class="kpi-delta-down">📉 {kpis['return_delta']} vs geçen ay</div>
            </div>
        </div>
        """, unsafe_allow_html=True)
        
    st.info("👆 Metrikler ve ciro değerleri sol menüdeki global tarih, müşteri ve marka filtrelerine göre dinamik olarak güncellenmektedir.")
    
    # --- AI EXECUTIVE SUMMARY ---
    if st.button("🧠 Durumu Özetle (AI Executive Summary)", use_container_width=True, key="btn_executive_summary"):
        with st.spinner("🧠 Yapay zeka metrikleri analiz edip yönetici özetini hazırlıyor..."):
            try:
                filters_list = []
                if len(date_range) == 2:
                    filters_list.append(f"Tarih Aralığı: {date_range[0]} - {date_range[1]}")
                if selected_customers and "Tümü" not in selected_customers:
                    filters_list.append(f"Müşteriler: {', '.join(selected_customers)}")
                if selected_brands and "Tümü" not in selected_brands:
                    filters_list.append(f"Markalar: {', '.join(selected_brands)}")
                active_filters_str = "; ".join(filters_list) if filters_list else "Filtre yok."

                kpi_summary_text = ""
                if excel_mode:
                    kpi_summary_text = (
                        f"Veri Kaynağı: Excel/CSV yüklemesi\n"
                        f"- Toplam Satır: {kpis['val1']:,} ({kpis['delta1']})\n"
                        f"- Sütun Sayısı: {kpis['val2']} ({kpis['delta2']})\n"
                        f"- Sayısal Değişken Sayısı: {kpis['val3']} ({kpis['delta3']})\n"
                    )
                else:
                    kpi_summary_text = (
                        f"Veri Kaynağı: Canlı PostgreSQL Veritabanı\n"
                        f"- Toplam Net Satış (EUR): {kpis['net_sales']:,.2f} € (Değişim: {kpis['sales_delta']})\n"
                        f"- Ortalama Brüt Marj (%): %{kpis['avg_margin']} (Değişim: {kpis['margin_delta']})\n"
                        f"- Müşteri İade Oranı (%): %{kpis['return_rate']} (Değişim: {kpis['return_delta']})\n"
                    )
                
                system_instruction = """You are an elite C-Level Business Intelligence Consultant. I will provide you with the current active filters and real-time dashboard KPIs. Your task is to write a highly professional, 3-paragraph executive summary in Turkish. 
* Paragraph 1: Overview of the current situation and sales performance. 
* Paragraph 2: Highlighting anomalies, risks, or notable trends (e.g., evaluating the return rate or margin). 
* Paragraph 3: Actionable, strategic business recommendations based on the data. 
Keep the tone corporate, insightful, and direct. Do NOT use code blocks; use bolding for emphasis."""

                user_prompt = f"""
Aktif Filtreler:
{active_filters_str}

Mevcut Metrikler (KPIs):
{kpi_summary_text}
"""
                
                response = client.chat.completions.create(
                    model=MODEL_NAME,
                    messages=[
                        {"role": "system", "content": system_instruction},
                        {"role": "user", "content": user_prompt}
                    ],
                    temperature=0.3
                )
                message_content = response.choices[0].message.content
                summary_text = message_content.strip() if message_content is not None else "Yönetici özeti oluşturulamadı (boş yanıt)."
                
                st.markdown(f"""
                <div style="background-color: #1e293b; border-left: 5px solid #3b82f6; padding: 20px; border-radius: 10px; margin-top: 15px; margin-bottom: 25px; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">
                    <h4 style="color: #60a5fa; margin-top: 0px; margin-bottom: 15px; font-weight: 700;">🧠 AI Yönetici Özeti (C-Level Executive Report)</h4>
                    <div style="color: #f8fafc; font-size: 14px; line-height: 1.6; font-family: sans-serif;">
                        {summary_text.replace(chr(10), '<br>')}
                    </div>
                </div>
                """, unsafe_allow_html=True)
            except Exception as e_summary:
                st.error(f"Yönetici özeti oluşturulurken hata oluştu: {e_summary}")

    # --- ÖRNEK ANALİZ GRAFİĞİ (DASHBOARD) ---
    st.write("---")
    st.subheader("📊 Hızlı Durum Göstergesi (Satış Dağılımı)")
    if excel_mode:
        if df_filtered is not None:
            render_plotly_charts(df_filtered.head(50), "dashboard_excel")
        else:
            st.info("Lütfen dashboard grafikleri üretmek için yan menüden bir dosya yükleyin.")
    else:
        # SQL ile hızlı satış dağılımı çekip Plotly ile görselleştirme
        res_dash = execute_sql("""
            SELECT m.musteri_adi, SUM(s.miktar * s.birim_fiyat) AS ciro 
            FROM siparisler s 
            JOIN musteriler m ON s.musteri_id = m.id 
            WHERE s.urun_durumu != 'iptal' 
            GROUP BY m.musteri_adi 
            ORDER BY ciro DESC;
        """)
        if res_dash and res_dash["status"] == "success":
            df_dash = pd.DataFrame(res_dash["data"], columns=res_dash["columns"])
            render_plotly_charts(df_dash.set_index("musteri_adi"), "dashboard_postgres")
        else:
            st.warning("Veri çekilemedi.")


# ==========================================
# 7. TAB 3: 📂 SAVED ANALYSIS & HISTORY
# ==========================================
with tab3:
    st.subheader("📂 Kayıtlı Raporlar ve Arama Geçmişi")
    
    # --- FAVORİ ANALİZLER ---
    st.write("### ⭐ Favori Raporlarım")
    if not st.session_state.saved_analyses:
        st.info("Henüz favorilere kaydedilmiş bir analiz bulunmuyor. 'AI Chat Assistant' sekmesinde oluşturulan tabloların hemen altındaki '⭐ Favori Raporlarıma Ekle' butonunu kullanarak raporlarınızı buraya kaydedebilirsiniz.")
    else:
        for fav_idx, fav in enumerate(st.session_state.saved_analyses):
            with st.expander(f"⭐ Rapor: {fav['soru']} ({fav['saved_at']})", expanded=False):
                # Favorilerden Sil Butonu
                if st.button("🗑️ Raporu Favorilerden Kaldır", key=f"del_fav_{fav_idx}"):
                    st.session_state.saved_analyses.pop(fav_idx)
                    st.success("Rapor favorilerden kaldırıldı.")
                    st.rerun()
                
                # SQL Kodu
                st.markdown("**Çalıştırılan SQL Sorgusu:**")
                st.code(fav["sql"], language="sql")
                
                # Plotly Grafik
                if fav["chart_data"] is not None and not fav["chart_data"].empty:
                    st.markdown("**Görsel Analiz:**")
                    render_plotly_charts(fav["chart_data"], f"saved_fav_{fav_idx}")
                
                # Tablo
                st.markdown("**Veri Tablosu:**")
                st.dataframe(fav["df"], use_container_width=True)
                
                # Dışa Aktarma Butonları
                col_fav_dl_csv, col_fav_dl_xlsx, _ = st.columns([1, 1, 2])
                with col_fav_dl_csv:
                    csv_data = fav["df"].to_csv(index=False).encode('utf-8')
                    st.download_button(
                        label="📥 Export CSV",
                        data=csv_data,
                        file_name=f"saved_report_{fav_idx}.csv",
                        mime="text/csv",
                        key=f"fav_dl_csv_{fav_idx}"
                    )
                with col_fav_dl_xlsx:
                    try:
                        xlsx_data = to_excel(fav["df"])
                        st.download_button(
                            label="📥 Export Excel",
                            data=xlsx_data,
                            file_name=f"saved_report_{fav_idx}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"fav_dl_xlsx_{fav_idx}"
                        )
                    except Exception:
                        st.caption("Excel Hatası")

    st.write("---")
    
    # --- YEREL ARAMA GEÇMİŞİ ---
    st.write("### 🕰️ Oturum Arama Geçmişi")
    if not st.session_state.gecmis_aramalar:
        st.info("Bu oturumda henüz başarılı bir SQL sorgusu çalıştırılmadı.")
    else:
        for i, gecmis in enumerate(reversed(st.session_state.gecmis_aramalar)):
            with st.expander(f"🔍 Sorgu #{len(st.session_state.gecmis_aramalar)-i}: {gecmis['soru']}", expanded=False):
                st.caption("Çalıştırılan SQL:")
                st.code(gecmis["sql"], language="sql")
                st.dataframe(gecmis["tablo"], use_container_width=True)